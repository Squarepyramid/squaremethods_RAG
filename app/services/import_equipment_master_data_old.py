"""
SquareMethods - Equipment Hierarchy Import Service
===================================================
Place this file at: app/services/import_equipment_master_data.py

Parses the Equipment Hierarchy Excel template (one sheet, one row
per equipment) and creates locations, equipment_types, and equipment
directly in the database. No async ingestion queue -- runs
synchronously end to end, called from a FastAPI endpoint.

Skip-and-report, not all-or-nothing: a row with a problem (duplicate
reference_code, missing required field, a malformed location path)
is skipped and reported, it does not block the rest of the file from
importing. Every row runs inside its own SAVEPOINT so an unexpected
DB error on one row only rolls back that row. Everything that
succeeds across the whole file is committed together at the end.

Template shape: any number of leading "Location Name" columns
(depth is however many are present, not fixed), followed by
"Equipment Name *", "Equipment Type Name *", "Reference Code *",
"Notes".

Location path rules:
  - Read location columns left to right. The first blank column
    ends the path -- nothing after it is read.
  - Consecutive columns repeating the same name are a genuine
    multi-level chain (e.g. a plant, line, and station that all
    happen to be named "Mixing Line"), not a stop signal.
  - A blank path (all location columns empty) means the equipment
    has no location.
  - A blank column followed by a filled one is a gap -- that row is
    skipped and reported rather than guessed at.
  - The same name can appear in different branches or at different
    depths across rows and refers to different location nodes each
    time -- identity is the full path (name + its parent chain),
    never the name alone. Locations are resolved/created level by
    level, matched on (company_id, name, parent_location_id).
  - A location is only created if it is the ancestor of at least
    one equipment row; there is no separate "define an empty
    location" step in this format.
"""

import io
import uuid
import logging
from typing import Optional

from openpyxl import load_workbook

from app.utils.db import get_db_connection

log = logging.getLogger(__name__)

SHEET_NAME = "Equipment Hierarchy"
EQUIPMENT_HEADERS = ["Equipment Name *", "Equipment Type Name *", "Reference Code *", "Notes"]


# ── Excel parser ──────────────────────────────────────────────────────────────

def _detect_columns(header_row: tuple) -> dict:
    """
    Detects column layout dynamically: however many leading columns
    are headed "Location Name" are treated as location depth levels,
    followed by the fixed equipment columns (matched by header text,
    not position, so column order in EQUIPMENT_HEADERS doesn't need
    to be exact as long as the headers are present).
    """
    headers = [str(h).strip() if h is not None else "" for h in header_row]

    n_location_cols = 0
    for h in headers:
        if h == "Location Name":
            n_location_cols += 1
        else:
            break

    if n_location_cols == 0:
        raise ValueError(
            "No 'Location Name' columns found. The Equipment Hierarchy tab "
            "must start with one or more columns headed exactly 'Location Name'."
        )

    equipment_col_idx = {}
    for name in EQUIPMENT_HEADERS:
        if name not in headers:
            raise ValueError(f"Missing required column '{name}' in Equipment Hierarchy tab.")
        equipment_col_idx[name] = headers.index(name)

    return {
        "n_location_cols": n_location_cols,
        "equipment_col_idx": equipment_col_idx,
    }


def parse_excel(file_bytes: bytes) -> list[dict]:
    """
    Parses the Equipment Hierarchy sheet into a list of dicts:
    {
        "path": ["Toronto Operations", "Plant 1 - Toronto", ...],  # may be []
        "equipment_name": str,
        "type_name": str,
        "reference_code": str,
        "notes": str | None,
        "row_number": int,  # 1-indexed Excel row, for error messages
    }
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Missing required tab '{SHEET_NAME}' in uploaded file.")

    ws = wb[SHEET_NAME]
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        raise ValueError(f"'{SHEET_NAME}' tab is empty.")

    layout = _detect_columns(header_row)
    n_loc = layout["n_location_cols"]
    idx = layout["equipment_col_idx"]

    equipment = []
    for row_number, row in enumerate(rows_iter, start=2):
        if not any(cell for cell in row if cell is not None):
            continue

        name = row[idx["Equipment Name *"]]
        name = str(name).strip() if name is not None else ""
        if not name:
            continue

        # Build the location path: stop at the first blank column.
        # Flag (rather than silently guess at) any non-blank value
        # appearing after a blank, since that indicates a gap the
        # user likely didn't intend.
        path = []
        hit_blank = False
        gap_found = False
        for c in range(n_loc):
            val = row[c] if c < len(row) else None
            val = str(val).strip() if val is not None else ""
            if val:
                if hit_blank:
                    gap_found = True
                else:
                    path.append(val)
            else:
                hit_blank = True

        if gap_found:
            equipment.append({
                "path": [],
                "equipment_name": name,
                "type_name": "",
                "reference_code": "",
                "notes": None,
                "row_number": row_number,
                "parse_error": (
                    "location path has a blank column followed by a filled one. "
                    "Fill location columns left to right with no gaps."
                ),
            })
            continue

        type_name = row[idx["Equipment Type Name *"]]
        type_name = str(type_name).strip() if type_name is not None else ""

        reference_code = row[idx["Reference Code *"]]
        reference_code = str(reference_code).strip() if reference_code is not None else ""

        notes = row[idx["Notes"]] if idx["Notes"] < len(row) else None
        notes = str(notes).strip() if notes is not None else None

        equipment.append({
            "path": path,
            "equipment_name": name,
            "type_name": type_name,
            "reference_code": reference_code,
            "notes": notes,
            "row_number": row_number,
            "parse_error": None,
        })

    log.info(f"Parsed {len(equipment)} equipment rows ({n_loc} location columns detected)")
    return equipment


# ── Row-level checks (used per-row inside save_equipment, not import-blocking) ─

def row_issue(e: dict, existing_ref_codes: set) -> Optional[str]:
    """
    Returns a human-readable reason this row should be skipped, or
    None if the row is fine to insert. Checked per-row so one bad
    row never blocks the rest of the file.
    """
    if e.get("parse_error"):
        return e["parse_error"]
    if not e["reference_code"]:
        return "Reference Code is required."
    if e["reference_code"] in existing_ref_codes:
        return f"reference_code '{e['reference_code']}' already exists for this company."
    if not e["type_name"]:
        return "Equipment Type Name is required."
    if not e["equipment_name"]:
        return "Equipment Name is required."
    return None


# ── DB helpers ────────────────────────────────────────────────────────────────

def resolve_location_path(conn, path: list[str], company_id: str, cache: dict) -> Optional[str]:
    """
    Resolves a full location path level by level, matching each level
    on (company_id, name, parent_location_id) so the same name can
    exist at different depths or in different branches without
    colliding. `cache` maps a full path tuple (up to and including
    that level) -> location id, shared across calls within one import
    so a repeated prefix across rows only hits the DB once.

    Returns the id of the deepest (final) location in the path, or
    None if path is empty.
    """
    parent_id = None
    prefix: tuple = ()

    for name in path:
        prefix = prefix + (name,)
        if prefix in cache:
            parent_id = cache[prefix]
            continue

        with conn.cursor() as cur:
            if parent_id is None:
                cur.execute("""
                    SELECT id FROM locations
                    WHERE company_id = %s::uuid
                    AND parent_location_id IS NULL
                    AND lower(name) = %s
                """, (company_id, name.lower()))
            else:
                cur.execute("""
                    SELECT id FROM locations
                    WHERE company_id = %s::uuid
                    AND parent_location_id = %s::uuid
                    AND lower(name) = %s
                """, (company_id, parent_id, name.lower()))
            row = cur.fetchone()

        if row:
            loc_id = str(row["id"])
        else:
            loc_id = str(uuid.uuid4())
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO locations
                        (id, company_id, name, parent_location_id, created_at, updated_at)
                    VALUES
                        (%s::uuid, %s::uuid, %s, %s::uuid, NOW(), NOW())
                """, (loc_id, company_id, name, parent_id))

        cache[prefix] = loc_id
        parent_id = loc_id

    return parent_id


def resolve_or_create_equipment_type(conn, type_name: str, company_id: str, type_id_map: dict) -> str:
    """
    Resolves an equipment type name to an id, creating it if needed.
    Checks the company's existing equipment_types first, then
    equipment_type_defaults (to link source_default_id), then
    creates a fresh company-scoped type as a last resort.
    """
    key = type_name.lower()
    if key in type_id_map:
        return type_id_map[key]

    with conn.cursor() as cur:
        cur.execute("""
            SELECT id FROM equipment_types
            WHERE company_id = %s::uuid AND lower(name) = %s
        """, (company_id, key))
        row = cur.fetchone()

    if row:
        type_id_map[key] = str(row["id"])
        return type_id_map[key]

    with conn.cursor() as cur:
        cur.execute(
            "SELECT id FROM equipment_type_defaults WHERE lower(name) = %s",
            (key,),
        )
        default_row = cur.fetchone()

    source_default_id = str(default_row["id"]) if default_row else None
    new_id = str(uuid.uuid4())
    with conn.cursor() as cur:
        cur.execute("""
            INSERT INTO equipment_types
                (id, company_id, name, source_default_id, created_at, updated_at)
            VALUES
                (%s::uuid, %s::uuid, %s, %s, NOW(), NOW())
        """, (new_id, company_id, type_name, source_default_id))

    type_id_map[key] = new_id
    return new_id


def save_equipment(conn, equipment: list[dict], company_id: str) -> dict:
    """
    Resolves each row's location path and equipment type, then inserts
    equipment -- skipping (not aborting) any row with a problem, such
    as a duplicate reference_code or a missing required field.

    Each row runs inside its own SAVEPOINT so a failure on one row
    (a constraint violation, an unexpected DB error, etc.) only rolls
    back that row; every other row in the file still gets processed
    and the successful ones are committed at the end.

    Returns {"created": [...], "skipped": [...]}.
    """
    with conn.cursor() as cur:
        cur.execute(
            "SELECT reference_code FROM equipment WHERE company_id = %s::uuid",
            (company_id,),
        )
        existing_ref_codes = {row["reference_code"] for row in cur.fetchall()}

    location_cache: dict = {}
    type_id_map: dict = {}
    created = []
    skipped = []

    for i, e in enumerate(equipment):
        savepoint = f"row_{i}"
        issue = row_issue(e, existing_ref_codes)
        if issue:
            skipped.append({
                "row_number": e["row_number"],
                "name": e["equipment_name"],
                "reference_code": e["reference_code"] or None,
                "reason": issue,
            })
            continue

        with conn.cursor() as cur:
            cur.execute(f"SAVEPOINT {savepoint}")

        try:
            location_id = resolve_location_path(conn, e["path"], company_id, location_cache)
            equipment_type_id = resolve_or_create_equipment_type(
                conn, e["type_name"], company_id, type_id_map
            )

            new_id = str(uuid.uuid4())
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO equipment
                        (id, company_id, equipment_type_id, location_id, name,
                         reference_code, notes, status, created_at, updated_at)
                    VALUES
                        (%s::uuid, %s::uuid, %s::uuid, %s::uuid, %s,
                         %s, %s, 'draft', NOW(), NOW())
                """, (
                    new_id, company_id, equipment_type_id, location_id,
                    e["equipment_name"], e["reference_code"], e["notes"],
                ))

            with conn.cursor() as cur:
                cur.execute(f"RELEASE SAVEPOINT {savepoint}")

            existing_ref_codes.add(e["reference_code"])
            created.append({
                "equipment_id": new_id,
                "name": e["equipment_name"],
                "reference_code": e["reference_code"],
                "location_path": " > ".join(e["path"]) if e["path"] else None,
            })

        except Exception as row_err:
            with conn.cursor() as cur:
                cur.execute(f"ROLLBACK TO SAVEPOINT {savepoint}")
            log.warning(f"Row {e['row_number']} failed and was skipped: {row_err}")
            skipped.append({
                "row_number": e["row_number"],
                "name": e["equipment_name"],
                "reference_code": e["reference_code"] or None,
                "reason": str(row_err),
            })

    log.info(
        f"{len(created)} equipment created, {len(skipped)} skipped, "
        f"{len(location_cache)} location nodes resolved for company {company_id}"
    )
    return {"created": created, "skipped": skipped}


# ── Main entry point ──────────────────────────────────────────────────────────

def ingest(file_bytes: bytes, company_id: str, created_by: str) -> dict:
    """
    Full import pipeline called from the FastAPI endpoint.
    1. Parse the Equipment Hierarchy sheet (dynamic location depth)
    2. Resolve/create each row's location path (name+parent identity)
       and equipment type, then insert equipment -- skipping any row
       with a problem (duplicate reference_code, missing required
       field, unresolvable data) rather than aborting the whole file
    3. Commit everything that succeeded as one transaction
    """
    equipment = parse_excel(file_bytes)

    if not equipment:
        raise ValueError(
            "No equipment rows found in the uploaded file. Check that the "
            "'Equipment Hierarchy' tab has data below the header row."
        )

    conn = get_db_connection()

    try:
        result = save_equipment(conn, equipment, company_id)
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    created = result["created"]
    skipped = result["skipped"]

    log.info(
        f"Import complete. {len(created)} equipment created, {len(skipped)} skipped "
        f"for company {company_id} (imported by {created_by})"
    )

    return {
        "company_id": company_id,
        "equipment_created": len(created),
        "equipment_skipped": len(skipped),
        "equipment": created,
        "skipped": skipped,
    }