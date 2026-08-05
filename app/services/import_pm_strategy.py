"""
SquareMethods - PM Strategy Import Service
==========================================
Place this file at: app/services/import_pm_strategy.py

Parses an edited PM strategy Excel file (the output of
generate_pm_strategy.py after reviewer edits) and creates
job aids and procedure steps in the database.

One job aid is created per PM type block that contains rows.
Job aid titles are prefixed with the equipment name so that
job aids for different equipment are distinguishable, e.g.
  "Wrapper 4 - PM2 - Lubrication"

Each job aid's image is set to the equipment's own image
(job_aids.image = equipment.image), so the job aid card shows
a photo of the equipment it belongs to.

Steps with an Image column value get that URL stored in the
procedures.image column. Steps with a blank Image column get
image = NULL.

PM type blocks are identified by header rows matching:
  "PM1 - Inspection", "PM2 - Lubrication", etc.
"""

import io
import re
import uuid
import logging
from typing import Optional

from openpyxl import load_workbook

from app.utils.db import get_db_connection

log = logging.getLogger(__name__)

PM_HEADER_PATTERN = re.compile(
    r"^(PM[1-9])\s*[-–]\s*(.+)$", re.IGNORECASE
)

CATEGORY_MAP = {
    "PM1": "Inspection",
    "PM2": "Lubrication",
    "PM3": "Calibration",
    "PM4": "Replacements",
    "PM5": "Overhaul",
    "PM6": "Condition Monitoring",
    "PM7": "Cleaning",
    "PM8": "Safety Inspection",
    "PM9": "Software Back-up",
}

# Column positions (1-indexed, matching the Excel template)
COL_OPERATION   = 1
COL_TITLE       = 2
COL_FREQUENCY   = 3
COL_HRS         = 4
COL_COMPONENT   = 8
COL_INSTRUCTION = 9
COL_FAILURE     = 10
COL_IMAGE       = 11


# ── Excel parser ──────────────────────────────────────────────────────────────

def parse_excel(file_bytes: bytes) -> list[dict]:
    """
    Parse the PM strategy Excel into a list of job aid blocks.

    Returns a list of dicts:
    {
        "pm_code":   "PM2",
        "pm_name":   "Lubrication",
        "category":  "Lubrication",
        "steps":     [ { step fields... }, ... ]
    }

    Skips PM type blocks that have no data rows.
    Skips the column header row (identified by "Operation" in col 1).
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    blocks       = []
    current_block = None

    for row in ws.iter_rows(values_only=True):
        # Skip completely empty rows
        if not any(cell for cell in row if cell is not None):
            continue

        first_cell = str(row[0]).strip() if row[0] is not None else ""

        # Detect PM type header row e.g. "PM2 - Lubrication"
        match = PM_HEADER_PATTERN.match(first_cell)
        if match:
            if current_block and current_block["steps"]:
                blocks.append(current_block)
            pm_code      = match.group(1).upper()
            pm_name      = CATEGORY_MAP.get(pm_code, match.group(2).strip())
            current_block = {
                "pm_code":  pm_code,
                "pm_name":  pm_name,
                "category": pm_name,
                "steps":    [],
            }
            continue

        # Skip column header row
        if first_cell.lower().startswith("operation") and current_block is None:
            continue
        if first_cell == "Operation":
            continue

        # Skip rows that are column headers (second row in each block)
        if first_cell in ("Operation", "Task List Description", "Frequency"):
            continue

        # Data row -- must be inside a block and have a step title
        if current_block is None:
            continue

        title = str(row[COL_TITLE - 1]).strip() if row[COL_TITLE - 1] is not None else ""
        if not title or title.lower() in ("none", "task list description"):
            continue

        # Parse hours -- may be blank or non-numeric
        hrs_raw = row[COL_HRS - 1]
        try:
            hrs = float(hrs_raw) if hrs_raw is not None else None
        except (ValueError, TypeError):
            hrs = None

        # Estimated duration in minutes for this step
        duration_minutes = round(hrs * 60) if hrs else None

        # Failure modes -- stored as precautions array
        failure_raw  = row[COL_FAILURE - 1]
        failure_text = str(failure_raw).strip() if failure_raw is not None else ""
        precautions  = (
            [f.strip() for f in failure_text.split(",") if f.strip()]
            if failure_text and failure_text.lower() != "none"
            else []
        )

        # Image URL -- blank becomes None
        image_raw = row[COL_IMAGE - 1]
        image_url = str(image_raw).strip() if image_raw is not None else ""
        image_url = image_url if image_url and image_url.lower() != "none" else None

        instruction_raw = row[COL_INSTRUCTION - 1]
        instruction     = str(instruction_raw).strip() if instruction_raw is not None else ""
        instruction     = instruction if instruction and instruction.lower() != "none" else ""

        component_raw = row[COL_COMPONENT - 1]
        component     = str(component_raw).strip() if component_raw is not None else ""

        frequency_raw = row[COL_FREQUENCY - 1]
        frequency     = str(frequency_raw).strip() if frequency_raw is not None else ""

        current_block["steps"].append({
            "title":            title,
            "instruction":      instruction,
            "component":        component,
            "frequency":        frequency,
            "hrs":              hrs,
            "duration_minutes": duration_minutes,
            "precautions":      precautions,
            "image_url":        image_url,
        })

    # Flush last block
    if current_block and current_block["steps"]:
        blocks.append(current_block)

    log.info(f"Parsed {len(blocks)} PM type blocks from Excel")
    return blocks


# ── Equipment lookup ──────────────────────────────────────────────────────────

def fetch_equipment(conn, equipment_id: str, company_id: str) -> dict:
    """
    Look up the equipment name and image so job aid titles can be prefixed
    with the equipment name (e.g. "Wrapper 4 - PM2 - Lubrication") and the
    job aid's image can be set to a photo of the equipment.
    Falls back to a generic name and no image if the equipment cannot be found.
    """
    with conn.cursor() as cur:
        cur.execute("""
            SELECT name, image FROM equipment
            WHERE id = %s::uuid
            AND company_id = %s::uuid
        """, (equipment_id, company_id))
        row = cur.fetchone()

    if row and row["name"]:
        return {"name": row["name"], "image": row.get("image")}

    log.warning(f"Equipment not found for {equipment_id}, using fallback name and no image")
    return {"name": "Equipment", "image": None}


# ── Slug helper ───────────────────────────────────────────────────────────────

def make_slug(title: str, unique_id: str) -> str:
    import re as _re
    base   = _re.sub(r"[^a-z0-9]+", "-", title.lower()).strip("-")
    suffix = unique_id[:8]
    return f"{base}-{suffix}"


# ── DB write ──────────────────────────────────────────────────────────────────

def save_block(
    conn,
    block: dict,
    equipment_id: str,
    equipment_name: str,
    equipment_image: Optional[str],
    company_id: str,
    created_by: str,
) -> str:
    """
    Insert one job aid and its procedure steps for a PM type block.
    Title is prefixed with the equipment name, e.g.
      "Wrapper 4 - PM2 - Lubrication"
    The job aid's image is set to the equipment's own image.
    Returns the job_aid_id.
    """
    job_aid_id = str(uuid.uuid4())
    title      = f"{equipment_name} - {block['pm_code']} - {block['pm_name']}"
    slug       = make_slug(title, job_aid_id)

    # estimated_duration = sum of all step durations in minutes
    total_minutes = sum(
        s["duration_minutes"] for s in block["steps"] if s["duration_minutes"]
    ) or None

    with conn.cursor() as cur:

        # 1. Insert job aid (image comes from the equipment, not the Excel)
        cur.execute("""
            INSERT INTO job_aids
                (id, company_id, title, slug, instruction, status,
                 estimated_duration, category, image, created_by,
                 view_count, scan_count, created_at, updated_at)
            VALUES
                (%s::uuid, %s::uuid, %s, %s, %s, 'draft',
                 %s, %s, %s, %s::uuid,
                 0, 0, NOW(), NOW())
        """, (
            job_aid_id,
            company_id,
            title,
            slug,
            f"Imported {block['pm_name']} job aid for {equipment_name}",
            total_minutes,
            block["category"],
            equipment_image,
            created_by,
        ))

        # 2. Insert procedure steps
        for step_num, step in enumerate(block["steps"], 1):
            cur.execute("""
                INSERT INTO procedures
                    (id, company_id, job_aid_id, title, step,
                     instruction, type, precautions, image,
                     created_at, updated_at)
                VALUES
                    (%s::uuid, %s::uuid, %s::uuid, %s, %s,
                     %s, 'procedure', %s, %s,
                     NOW(), NOW())
            """, (
                str(uuid.uuid4()),
                company_id,
                job_aid_id,
                step["title"],
                step_num,
                step["instruction"],
                step["precautions"],
                step["image_url"],
            ))

        # 3. Link job aid to equipment node
        cur.execute("""
            INSERT INTO job_aid_equipment
                (id, company_id, job_aid_id, equipment_id,
                 created_at, updated_at)
            VALUES
                (%s::uuid, %s::uuid, %s::uuid, %s::uuid,
                 NOW(), NOW())
        """, (
            str(uuid.uuid4()),
            company_id,
            job_aid_id,
            equipment_id,
        ))

    log.info(f"Saved job aid {job_aid_id} ({title}) with {len(block['steps'])} steps")
    return job_aid_id


# ── Main entry point ──────────────────────────────────────────────────────────

def ingest(
    file_bytes: bytes,
    equipment_id: str,
    company_id: str,
    created_by: str,
) -> dict:
    """
    Full import pipeline called from the FastAPI endpoint.
    1. Look up the equipment name and image for title prefixing / job aid image
    2. Parse Excel into PM type blocks
    3. Save each block as a job aid with procedure steps
    4. Return summary of created job aids
    """
    blocks = parse_excel(file_bytes)

    if not blocks:
        raise ValueError(
            "No valid PM strategy data found in the uploaded file. "
            "Ensure the file follows the PM strategy template format."
        )

    conn        = get_db_connection()
    created_ids = []

    try:
        equipment       = fetch_equipment(conn, equipment_id, company_id)
        equipment_name  = equipment["name"]
        equipment_image = equipment["image"]

        for block in blocks:
            job_aid_id = save_block(
                conn, block, equipment_id, equipment_name, equipment_image,
                company_id, created_by
            )
            created_ids.append({
                "job_aid_id": job_aid_id,
                "pm_code":    block["pm_code"],
                "pm_name":    block["pm_name"],
                "steps":      len(block["steps"]),
            })
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    log.info(f"Import complete. {len(created_ids)} job aids created for equipment {equipment_id} ({equipment_name})")

    return {
        "equipment_id": equipment_id,
        "equipment_name": equipment_name,
        "equipment_image": equipment_image,
        "job_aids_created": len(created_ids),
        "job_aids": created_ids,
    }