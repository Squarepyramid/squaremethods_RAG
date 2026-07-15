"""
SquareMethods - Equipment Hierarchy Import Service
===================================================
Place this file at: app/services/import_equipment_master_data.py

Parses an Excel workbook with up to two tabs and applies them to the
database in one transaction, structural changes first, then equipment:

  1. "Location Moves" (optional tab) -- explicit hierarchy
     restructuring: move an existing location to a new parent, or
     insert a new level directly above one. Processed first so the
     Equipment Hierarchy tab below sees the updated tree. This is
     deliberately explicit, not inferred from a changed path on an
     equipment row -- moving a location moves its entire subtree
     (every child location and every piece of equipment under it at
     any depth), so it needs its own unambiguous instruction rather
     than a guess based on a name that looks similar to something
     seen before.

  2. "Equipment Hierarchy" (required tab) -- one row per equipment.
     A row whose Reference Code already exists for this company is
     treated as an UPDATE (location, equipment type, name, notes),
     not a duplicate to skip -- this is what lets a single piece of
     equipment be repositioned by simply re-uploading it with a new
     location path. A row with a brand-new Reference Code is created.

Skip-and-report throughout, not all-or-nothing: a row or move
instruction with a problem (duplicate reference_code with no
resolvable change, missing required field, a malformed location
path, a path that can't be found) is skipped and reported, it does
not block the rest of the file. Every row/instruction runs inside
its own SAVEPOINT so an unexpected DB error on one only rolls back
that one. Everything that succeeds across the whole file is
committed together at the end.

Equipment Hierarchy tab shape: any number of leading "Location Name"
columns (depth is however many are present, not fixed), followed by
"Equipment Name *", "Equipment Type Name *", "Reference Code *",
"Notes".

Location path rules (Equipment Hierarchy tab):
  - Read location columns left to right. The first blank column
    ends the path -- nothing after it is read.
  - Consecutive columns repeating the same name are a genuine
    multi-level chain (e.g. a plant, line, and station that all
    happen to be named "Mixing Line"), not a stop signal.
  - A blank path (all location columns empty) means the equipment
    has no location.
  - A blank column followed by a filled one is a gap -- that row is
    skipped and reported rather than guessed at.
  - The same name can appear in different branches or at different
    depths across rows and refers to different location nodes each
    time -- identity is the full path (name + its parent chain),
    never the name alone. Locations are resolved/created level by
    level, matched on (company_id, name, parent_location_id).
  - A location is only created if it is the ancestor of at least
    one equipment row; there is no separate "define an empty
    location" step in this format.

Location Moves tab shape: "Action *" (Move | Insert Level),
"Existing Location Path *" (a single ' > '-separated string, e.g.
"Plant 1 - Toronto > Packaging Line 1"), "New Parent Path" (used for
Move; ' > '-separated, blank means top-level), "New Level Name"
(used for Insert Level).
"""

import io
import uuid
import logging
from typing import Optional

from openpyxl import load_workbook

from app.utils.db import get_db_connection
from app.services.manage_locations import move_location, insert_location_level

log = logging.getLogger(__name__)

SHEET_NAME = "Equipment Hierarchy"
SHEET_NAME_MOVES = "Location Moves"
EQUIPMENT_HEADERS = ["Equipment Name *", "Equipment Type Name *", "Reference Code *", "Notes"]
MOVE_HEADERS = ["Action *", "Existing Location Path *", "New Parent Path", "New Level Name"]


# ── Excel parser ──────────────────────────────────────────────────────────────

def _detect_columns(header_row: tuple) -> dict:
    """
    Detects column layout dynamically: however many leading columns
    are headed "Location Name" are treated as location depth levels,
    followed by the fixed equipment columns (matched by header text,
    not position, so column order in EQUIPMENT_HEADERS doesn't need
    to be exact as long as the headers are present).
    """
    headers = [str(h).strip() if h is not None else "" for h in header_row]

    n_location_cols = 0
    for h in headers:
        if h == "Location Name":
            n_location_cols += 1
        else:
            break

    if n_location_cols == 0:
        raise ValueError(
            "No 'Location Name' columns found. The Equipment Hierarchy tab "
            "must start with one or more columns headed exactly 'Location Name'."
        )

    equipment_col_idx = {}
    for name in EQUIPMENT_HEADERS:
        if name not in headers:
            raise ValueError(f"Missing required column '{name}' in Equipment Hierarchy tab.")
        equipment_col_idx[name] = headers.index(name)

    return {
        "n_location_cols": n_location_cols,
        "equipment_col_idx": equipment_col_idx,
    }


def parse_excel(file_bytes: bytes) -> list[dict]:
    """
    Parses the Equipment Hierarchy sheet into a list of dicts:
    {
        "path": ["Toronto Operations", "Plant 1 - Toronto", ...],  # may be []
        "equipment_name": str,
        "type_name": str,
        "reference_code": str,
        "notes": str | None,
        "row_number": int,  # 1-indexed Excel row, for error messages
    }
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Missing required tab '{SHEET_NAME}' in uploaded file.")

    ws = wb[SHEET_NAME]
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        raise ValueError(f"'{SHEET_NAME}' tab is empty.")

    layout = _detect_columns(header_row)
    n_loc = layout["n_location_cols"]
    idx = layout["equipment_col_idx"]

    equipment = []
    for row_number, row in enumerate(rows_iter, start=2):
        if not any(cell for cell in row if cell is not None):
            continue

        name = row[idx["Equipment Name *"]]
        name = str(name).strip() if name is not None else ""
        if not name:
            continue

        # Build the location path: stop at the first blank column.
        # Flag (rather than silently guess at) any non-blank value
        # appearing after a blank, since that indicates a gap the
        # user likely didn't intend.
        path = []
        hit_blank = False
        gap_found = False
        for c in range(n_loc):
            val = row[c] if c < len(row) else None
            val = str(val).strip() if val is not None else ""
            if val:
                if hit_blank:
                    gap_found = True
                else:
                    path.append(val)
            else:
                hit_blank = True

        if gap_found:
            equipment.append({
                "path": [],
                "equipment_name": name,
                "type_name": "",
                "reference_code": "",
                "notes": None,
                "row_number": row_number,
                "parse_error": (
                    "location path has a blank column followed by a filled one. "
                    "Fill location columns left to right with no gaps."
                ),
            })
            continue

        type_name = row[idx["Equipment Type Name *"]]
        type_name = str(type_name).strip() if type_name is not None else ""

        reference_code = row[idx["Reference Code *"]]
        reference_code = str(reference_code).strip() if reference_code is not None else ""

        notes = row[idx["Notes"]] if idx["Notes"] < len(row) else None
        notes = str(notes).strip() if notes is not None else None

        equipment.append({
            "path": path,
            "equipment_name": name,
            "type_name": type_name,
            "reference_code": reference_code,
            "notes": notes,
            "row_number": row_number,
            "parse_error": None,
        })

    log.info(f"Parsed {len(equipment)} equipment rows ({n_loc} location columns detected)")
    return equipment


# ── Location Moves parser (optional tab) ──────────────────────────────────────

def _split_path(raw: object) -> list[str]:
    if raw is None:
        return []
    text = str(raw).strip()
    if not text:
        return []
    return [p.strip() for p in text.split(">") if p.strip()]


def parse_location_moves(file_bytes: bytes) -> list[dict]:
    """
    Parses the optional "Location Moves" tab into a list of dicts:
    {
        "action": "move" | "insert_level",
        "existing_path": [...],
        "new_parent_path": [...] | None,   # used for "move"
        "new_level_name": str | None,      # used for "insert_level"
        "row_number": int,
    }
    Returns [] if the tab isn't present -- it's optional, most
    imports are pure equipment additions with no restructuring.
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    if SHEET_NAME_MOVES not in wb.sheetnames:
        return []

    ws = wb[SHEET_NAME_MOVES]
    all_rows = list(ws.iter_rows(values_only=True))

    header_row_idx = None
    for i, row in enumerate(all_rows):
        cells = [str(c).strip() if c is not None else "" for c in row]
        if "Action *" in cells:
            header_row_idx = i
            break

    if header_row_idx is None:
        return []

    header_row = all_rows[header_row_idx]
    data_rows = all_rows[header_row_idx + 1:]
    data_row_offset = header_row_idx + 2  # 1-indexed Excel row number of the first data row

    headers = [str(h).strip() if h is not None else "" for h in header_row]
    idx = {}
    for name in MOVE_HEADERS:
        if name not in headers:
            raise ValueError(f"Missing required column '{name}' in Location Moves tab.")
        idx[name] = headers.index(name)

    moves = []
    for offset, row in enumerate(data_rows):
        row_number = data_row_offset + offset
        if not any(cell for cell in row if cell is not None):
            continue

        action_raw = row[idx["Action *"]]
        action_raw = str(action_raw).strip().lower() if action_raw is not None else ""
        if not action_raw:
            continue

        if action_raw in ("move",):
            action = "move"
        elif action_raw in ("insert level", "insert_level", "insertlevel"):
            action = "insert_level"
        else:
            moves.append({
                "action": None,
                "existing_path": [],
                "new_parent_path": None,
                "new_level_name": None,
                "row_number": row_number,
                "parse_error": f"Unrecognized Action '{action_raw}'. Use 'Move' or 'Insert Level'.",
            })
            continue

        existing_path = _split_path(row[idx["Existing Location Path *"]])
        new_parent_path = _split_path(row[idx["New Parent Path"]]) if idx["New Parent Path"] < len(row) else []
        new_level_name_raw = row[idx["New Level Name"]] if idx["New Level Name"] < len(row) else None
        new_level_name = str(new_level_name_raw).strip() if new_level_name_raw is not None else ""

        moves.append({
            "action": action,
            "existing_path": existing_path,
            "new_parent_path": new_parent_path or None,
            "new_level_name": new_level_name or None,
            "row_number": row_number,
            "parse_error": None,
        })

    log.info(f"Parsed {len(moves)} location move instructions")
    return moves


def apply_location_moves(conn, moves: list[dict], company_id: str) -> dict:
    """
    Applies each Location Moves instruction inside its own SAVEPOINT,
    skipping (not aborting) any instruction with a problem -- an
    unrecognized action, a missing path, or a path that can't be
    found. Everything that succeeds is left in place for the
    Equipment Hierarchy tab to build on afterward.
    """
    applied = []
    skipped = []

    for i, m in enumerate(moves):
        savepoint = f"move_{i}"

        if m.get("parse_error"):
            skipped.append({"row_number": m["row_number"], "reason": m["parse_error"]})
            continue
        if not m["existing_path"]:
            skipped.append({"row_number": m["row_number"], "reason": "Existing Location Path is required."})
            continue
        if m["action"] == "insert_level" and not m["new_level_name"]:
            skipped.append({"row_number": m["row_number"], "reason": "New Level Name is required for Insert Level."})
            continue

        with conn.cursor() as cur:
            cur.execute(f"SAVEPOINT {savepoint}")

        try:
            if m["action"] == "move":
                result = move_location(conn, company_id, m["existing_path"], m["new_parent_path"])
            else:
                result = insert_location_level(conn, company_id, m["existing_path"], m["new_level_name"])

            with conn.cursor() as cur:
                cur.execute(f"RELEASE SAVEPOINT {savepoint}")

            applied.append({
                "row_number": m["row_number"],
                "action": m["action"],
                "existing_path": " > ".join(m["existing_path"]),
                "result": result,
            })

        except Exception as move_err:
            with conn.cursor() as cur:
                cur.execute(f"ROLLBACK TO SAVEPOINT {savepoint}")
            log.warning(f"Location move row {m['row_number']} failed and was skipped: {move_err}")
            skipped.append({"row_number": m["row_number"], "reason": str(move_err)})

    log.info(f"{len(applied)} location moves applied, {len(skipped)} skipped for company {company_id}")
    return {"applied": applied, "skipped": skipped}


# ── Row-level checks (used per-row inside save_equipment, not import-blocking) ─

def row_issue(e: dict) -> Optional[str]:
    """
    Returns a human-readable reason this row can't be processed at
    all (skip, regardless of insert vs update), or None if it's fine.
    A pre-existing reference_code is NOT a skip reason on its own --
    that's handled as an update in save_equipment.
    """
    if e.get("parse_error"):
        return e["parse_error"]
    if not e["reference_code"]:
        return "Reference Code is required."
    if not e["type_name"]:
        return "Equipment Type Name is required."
    if not e["equipment_name"]:
        return "Equipment Name is required."
    return None


# ── DB helpers ────────────────────────────────────────────────────────────────

def resolve_location_path(conn, path: list[str], company_id: str, cache: dict) -> Optional[str]:
    """
    Resolves a full location path level by level, matching each level
    on (company_id, name, parent_location_id) so the same name can
    exist at different depths or in different branches without
    colliding. `cache` maps a full path tuple (up to and including
    that level) -> location id, shared across calls within one import
    so a repeated prefix across rows only hits the DB once.

    Returns the id of the deepest (final) location in the path, or
    None if path is empty.
    """
    parent_id = None
    prefix: tuple = ()

    for name in path:
        prefix = prefix + (name,)
        if prefix in cache:
            parent_id = cache[prefix]
            continue

        with conn.cursor() as cur:
            if parent_id is None:
                cur.execute("""
                    SELECT id FROM locations
                    WHERE company_id = %s::uuid
                    AND parent_location_id IS NULL
                    AND lower(name) = %s
                """, (company_id, name.lower()))
            else:
                cur.execute("""
                    SELECT id FROM locations
                    WHERE company_id = %s::uuid
                    AND parent_location_id = %s::uuid
                    AND lower(name) = %s
                """, (company_id, parent_id, name.lower()))
            row = cur.fetchone()

        if row:
            loc_id = str(row["id"])
        else:
            loc_id = str(uuid.uuid4())
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO locations
                        (id, company_id, name, parent_location_id, created_at, updated_at)
                    VALUES
                        (%s::uuid, %s::uuid, %s, %s::uuid, NOW(), NOW())
                """, (loc_id, company_id, name, parent_id))

        cache[prefix] = loc_id
        parent_id = loc_id

    return parent_id


def resolve_or_create_equipment_type(conn, type_name: str, company_id: str, type_id_map: dict) -> str:
    """
    Resolves an equipment type name to an id, creating it if needed.
    Checks the company's existing equipment_types first, then
    equipment_type_defaults (to link source_default_id), then
    creates a fresh company-scoped type as a last resort.
    """
    key = type_name.lower()
    if key in type_id_map:
        return type_id_map[key]

    with conn.cursor() as cur:
        cur.execute("""
            SELECT id FROM equipment_types
            WHERE company_id = %s::uuid AND lower(name) = %s
        """, (company_id, key))
        row = cur.fetchone()

    if row:
        type_id_map[key] = str(row["id"])
        return type_id_map[key]

    with conn.cursor() as cur:
        cur.execute(
            "SELECT id FROM equipment_type_defaults WHERE lower(name) = %s",
            (key,),
        )
        default_row = cur.fetchone()

    source_default_id = str(default_row["id"]) if default_row else None
    new_id = str(uuid.uuid4())
    with conn.cursor() as cur:
        cur.execute("""
            INSERT INTO equipment_types
                (id, company_id, name, source_default_id, created_at, updated_at)
            VALUES
                (%s::uuid, %s::uuid, %s, %s, NOW(), NOW())
        """, (new_id, company_id, type_name, source_default_id))

    type_id_map[key] = new_id
    return new_id


def save_equipment(conn, equipment: list[dict], company_id: str) -> dict:
    """
    Resolves each row's location path and equipment type, then either
    inserts a new equipment row or -- if its Reference Code already
    exists for this company -- updates that existing row's location,
    type, name, and notes. This is what lets a single piece of
    equipment be repositioned by simply re-uploading it with a new
    location path, without touching anything else in the hierarchy.

    Rows with a missing required field or an unresolvable location
    path are skipped and reported rather than aborting the file.

    Each row runs inside its own SAVEPOINT so a failure on one row
    (a constraint violation, an unexpected DB error, etc.) only rolls
    back that row; every other row still gets processed.

    Returns {"created": [...], "updated": [...], "skipped": [...]}.
    """
    with conn.cursor() as cur:
        cur.execute(
            "SELECT id, reference_code FROM equipment WHERE company_id = %s::uuid",
            (company_id,),
        )
        existing_by_ref_code = {row["reference_code"]: str(row["id"]) for row in cur.fetchall()}

    location_cache: dict = {}
    type_id_map: dict = {}
    created = []
    updated = []
    skipped = []

    for i, e in enumerate(equipment):
        savepoint = f"row_{i}"
        issue = row_issue(e)
        if issue:
            skipped.append({
                "row_number": e["row_number"],
                "name": e["equipment_name"],
                "reference_code": e["reference_code"] or None,
                "reason": issue,
            })
            continue

        existing_equipment_id = existing_by_ref_code.get(e["reference_code"])

        with conn.cursor() as cur:
            cur.execute(f"SAVEPOINT {savepoint}")

        try:
            location_id = resolve_location_path(conn, e["path"], company_id, location_cache)
            equipment_type_id = resolve_or_create_equipment_type(
                conn, e["type_name"], company_id, type_id_map
            )

            if existing_equipment_id:
                with conn.cursor() as cur:
                    cur.execute("""
                        UPDATE equipment
                        SET equipment_type_id = %s::uuid,
                            location_id = %s::uuid,
                            name = %s,
                            notes = %s,
                            updated_at = NOW()
                        WHERE id = %s::uuid AND company_id = %s::uuid
                    """, (
                        equipment_type_id, location_id, e["equipment_name"], e["notes"],
                        existing_equipment_id, company_id,
                    ))

                with conn.cursor() as cur:
                    cur.execute(f"RELEASE SAVEPOINT {savepoint}")

                updated.append({
                    "equipment_id": existing_equipment_id,
                    "name": e["equipment_name"],
                    "reference_code": e["reference_code"],
                    "location_path": " > ".join(e["path"]) if e["path"] else None,
                })

            else:
                new_id = str(uuid.uuid4())
                with conn.cursor() as cur:
                    cur.execute("""
                        INSERT INTO equipment
                            (id, company_id, equipment_type_id, location_id, name,
                             reference_code, notes, status, created_at, updated_at)
                        VALUES
                            (%s::uuid, %s::uuid, %s::uuid, %s::uuid, %s,
                             %s, %s, 'draft', NOW(), NOW())
                    """, (
                        new_id, company_id, equipment_type_id, location_id,
                        e["equipment_name"], e["reference_code"], e["notes"],
                    ))

                with conn.cursor() as cur:
                    cur.execute(f"RELEASE SAVEPOINT {savepoint}")

                existing_by_ref_code[e["reference_code"]] = new_id
                created.append({
                    "equipment_id": new_id,
                    "name": e["equipment_name"],
                    "reference_code": e["reference_code"],
                    "location_path": " > ".join(e["path"]) if e["path"] else None,
                })

        except Exception as row_err:
            with conn.cursor() as cur:
                cur.execute(f"ROLLBACK TO SAVEPOINT {savepoint}")
            log.warning(f"Row {e['row_number']} failed and was skipped: {row_err}")
            skipped.append({
                "row_number": e["row_number"],
                "name": e["equipment_name"],
                "reference_code": e["reference_code"] or None,
                "reason": str(row_err),
            })

    log.info(
        f"{len(created)} equipment created, {len(updated)} updated, {len(skipped)} skipped, "
        f"{len(location_cache)} location nodes resolved for company {company_id}"
    )
    return {"created": created, "updated": updated, "skipped": skipped}


# ── Main entry point ──────────────────────────────────────────────────────────

def ingest(file_bytes: bytes, company_id: str, created_by: str) -> dict:
    """
    Full import pipeline called from the FastAPI endpoint.
    1. Parse and apply "Location Moves" (optional tab) -- explicit
       restructuring, processed first so the hierarchy is in its
       final shape before equipment is resolved against it
    2. Parse the "Equipment Hierarchy" tab (dynamic location depth)
    3. Resolve/create each row's location path and equipment type,
       then insert new equipment or update existing equipment
       (matched by Reference Code) -- skipping any row/instruction
       with a problem rather than aborting the whole file
    4. Commit everything that succeeded as one transaction
    """
    moves = parse_location_moves(file_bytes)
    equipment = parse_excel(file_bytes)

    if not equipment:
        raise ValueError(
            "No equipment rows found in the uploaded file. Check that the "
            "'Equipment Hierarchy' tab has data below the header row."
        )

    conn = get_db_connection()

    try:
        move_result = apply_location_moves(conn, moves, company_id) if moves else {"applied": [], "skipped": []}
        equipment_result = save_equipment(conn, equipment, company_id)
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    created = equipment_result["created"]
    updated = equipment_result["updated"]
    skipped = equipment_result["skipped"]

    log.info(
        f"Import complete for company {company_id} (imported by {created_by}): "
        f"{len(move_result['applied'])} location moves applied, "
        f"{len(created)} equipment created, {len(updated)} updated, {len(skipped)} skipped"
    )

    return {
        "company_id": company_id,
        "location_moves_applied": len(move_result["applied"]),
        "location_moves_skipped": len(move_result["skipped"]),
        "location_moves": move_result,
        "equipment_created": len(created),
        "equipment_updated": len(updated),
        "equipment_skipped": len(skipped),
        "equipment": {"created": created, "updated": updated, "skipped": skipped},
    }