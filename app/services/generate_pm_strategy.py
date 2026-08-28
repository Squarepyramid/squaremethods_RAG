"""
SquareMethods - PM Strategy Generation Service
==============================================
Place this file at: app/services/generate_pm_strategy.py

Pulls the manual chunks most relevant to each PM type for an equipment
node, then runs ten parallel async Claude calls (Working Principle +
nine PM types) to extract and structure maintenance tasks into a
downloadable Excel file.

*** IMAGE MATCHING IS CURRENTLY DISABLED (IMAGE_MATCHING_ENABLED = False) ***
This mirrors ingest_document.py's IMAGE_EXTRACTION_ENABLED flag -- since
ingest no longer writes to equipment_manual_images, there's nothing for
this step to match against anyway, so this flag just skips the (now
pointless) DB round-trip rather than querying an always-empty table on
every generation run. The "Image" column stays in the Excel output
(the import format expects it) but every cell comes back blank while
this is off. All the matching logic (match_manual_image,
resolve_component_image_url, attach_images_to_steps,
fetch_manual_images_for_equipment) is untouched and ready to go the
moment both flags are flipped back to True.

The Excel output matches the PM_strategy.xlsx format exactly so
the reviewer can fill gaps, add image URLs, then upload it via
the import endpoint. Use generate_with_filename() rather than calling
generate() directly if you want the file to come with a real,
human-readable name (e.g. "PM_Strategy_Compressor_1_A102_2026-08-05.xlsx")
instead of having to build one yourself -- see that function's docstring.

PM Types:
  WP  - Working Principle (step-by-step operating procedure, not a
        maintenance task -- placed first as the foundational section
        the rest of the strategies build on)
  PM1 - Inspection
  PM2 - Lubrication
  PM3 - Calibration
  PM4 - Replacements
  PM5 - Overhaul
  PM6 - Condition Monitoring
  PM7 - Cleaning
  PM8 - Safety Inspection
  PM9 - Software Back-up

CHANGE LOG (2026-08-28 revision -- reliability/consistency pass):
  This revision addresses three production incidents traced back to this
  file: a Bedrock ThrottlingException mid-generation (job codes mlf001/
  mlf003), a WP category silently returning 0 steps because Claude
  returned key-value text instead of JSON, and a reproducible PM8
  json.JSONDecodeError. All three trace back to the same root design
  choice -- asking Claude for free-text JSON and then hoping
  raw_text.find("[") / rfind("]") / json.loads() succeeds -- plus a
  missing temperature setting that made every rerun of the same manual
  produce a different number of populated categories. Also: this
  deployment is multi-tenant, and Bedrock's on-demand quota is account
  + region + model level, not per-tenant, so one tenant's burst can
  throttle another's job with zero visibility into which tenant did it.
  Changes below:

  - MODEL: BEDROCK_MODEL now defaults to a Haiku 4.5-class model (see
    the constant below) instead of claude-3-haiku-20240307, which is
    now legacy on Bedrock. VERIFY THE EXACT MODEL ID IN YOUR OWN BEDROCK
    CONSOLE / MODEL CATALOG FOR YOUR REGION BEFORE DEPLOYING -- the ID
    below is correct as of this writing but Anthropic/AWS model IDs and
    version suffixes do change, and this is exactly the kind of typo
    that fails loudly (ValidationException on an unknown model ID) so
    it's a cheap thing to double check once rather than trust blindly.
    For maximum extraction quality on customer-facing samples (more
    reliable category classification on dense/ambiguous manuals, at
    higher per-call cost/latency), swap to a Sonnet-tier model instead
    via the BEDROCK_MODEL_ID env var -- no code change needed either way.

  - STRUCTURED OUTPUT VIA FORCED TOOL USE: replaced the "return ONLY a
    valid JSON array" free-text instruction + raw_text.find("[")/
    rfind("]")/json.loads() extraction with Anthropic's tool-use
    mechanism: a single EXTRACTION_TOOL schema is sent on every call
    with tool_choice forcing Claude to call it. The model's structured
    arguments come back in response["content"][*]["input"] as an
    ALREADY-PARSED object -- no bracket-finding, no json.loads() on
    model-generated text, no way for a stray unescaped quote in an
    instruction field to corrupt the whole array, and no way for the
    model to "forget" and return prose instead (that's what caused the
    WP category's silent 0-step failures). This works on every Claude
    model on Bedrock, not just the newest ones -- it's not tied to the
    model swap above.

  - TEMPERATURE: added temperature=0 to every call. Previously unset
    (Bedrock default ~1.0), which is what let the exact same manual
    produce 5/10 populated categories on one run and 7/10 on the next.
    This is an extraction task ("find every task matching category X in
    this manual"), not a creative one -- there's no reason to want
    sampling variance here. This alone won't make output byte-identical
    run to run (floating-point non-associativity in batched inference
    means even temperature=0 isn't perfectly deterministic) but it
    removes the dominant source of variance.

  - stop_reason NOW LOGGED: previously nothing inspected
    raw.get("stop_reason"), so a response truncated by hitting MAX_TOKENS
    was indistinguishable in the logs from a clean parse failure. Now
    logged explicitly per call, and MAX_TOKENS raised from 4096 to 8192
    for headroom on categories that can legitimately produce 20+
    multi-step tasks (PM1, PM8 have both done this in production).

  - PER-CATEGORY STATUS ("ok" / "empty" / "error"): call_claude_for_pm_type
    now returns a status alongside the steps list instead of collapsing
    every failure mode (throttling, timeout, malformed response, a
    legitimately empty category) into an indistinguishable []. A category
    that errored is marked visibly in the Excel output instead of looking
    identical to "the manual doesn't cover this" -- see build_excel.

  - RETRY CONFIG: BEDROCK_CLIENT_CONFIG retries raised from
    {"max_attempts": 2, "mode": "standard"} to
    {"total_max_attempts": 6, "mode": "standard"}, matching AWS's current
    documented guidance for handling Bedrock throttling
    (see: Scaling and throughput best practices, Amazon Bedrock docs).

  - IN-PROCESS CONCURRENCY CAP: a semaphore now caps how many of this
    job's own 10 category calls hit Bedrock at the same instant
    (BEDROCK_CALL_CONCURRENCY, default 5) instead of firing all 10
    simultaneously. NOTE: this only smooths out the burst a single job
    creates by itself -- it does NOT coordinate across separate Lambda
    invocations or tenants. The actual cross-tenant throttling problem
    (tenant A's burst throttling tenant B's job) needs an infra-level
    fix outside this file: a Lambda reserved-concurrency cap, a shared
    token bucket (DynamoDB/ElastiCache), and/or Bedrock cross-region
    inference to raise the effective ceiling. This semaphore is a
    partial mitigation, not the full fix.

  - TENANT-SCOPED LOGGING: company_id is now included in every log line
    in this module. Previously it was only present in SQL query
    parameters, not in the log messages themselves, which made it
    impossible to filter "show me only this tenant's generation logs"
    without cross-referencing DB queries by hand.

  Everything below this point (retrieval, image matching, Excel
  building, filename generation) is UNCHANGED from the prior revision
  except for the log-line company_id additions noted above.

CHANGE LOG ADDENDUM (same day, second pass): this file originally built
its own boto3 bedrock-runtime client and its own request body directly
(duplicating model ID, retry Config, and response logging that also
needed to exist, separately, in app/services/bedrock_client.py for
/chat, ingest_document.py, and generate_job_aid.py). Reworked to call
app.services.bedrock_client.call_claude() instead: BEDROCK_MODEL,
BEDROCK_CLIENT_CONFIG, and the bedrock client construction that used to
live in this file are gone (see bedrock_client.py's own changelog for
what replaced them there); call_claude_for_pm_type() no longer takes a
`bedrock` client parameter, and the invoke_model/body-read/stop_reason
logging that used to be inline here now happens once, centrally, inside
call_claude(). Net effect: the retry/timeout hardening and structured
tool-use call from this revision now protect every Bedrock call site in
the app, not just PM strategy generation.

CHANGE LOG ADDENDUM (same day, third pass -- optimize for reliability, not
speed): explicit direction that this Lambda has up to 15 minutes to work
with, generation isn't latency-sensitive, and every prior fix should lean
toward resilience over throughput. Changes:
  - BEDROCK_CALL_CONCURRENCY lowered from 5 to 3 (env-overridable): a
    smaller burst per job is gentler on the shared Bedrock quota for a
    modest wall-clock cost this workload can easily absorb.
  - NEW: an application-level retry loop per category (PM_CATEGORY_MAX_ATTEMPTS,
    default 3, with linear backoff via PM_CATEGORY_RETRY_BACKOFF_SECONDS).
    Previously a category that got a usable HTTP response but an unusable
    one (no tool_use block, an anomaly worth retrying rather than the
    empty-array case) had zero retries and was permanently lost -- the
    only retries in the whole system were botocore's, and only for
    network-level failures. This is layered on top of those, not a
    replacement for them.
  - NEW: GENERATION_DEADLINE_SECONDS, a shared time.monotonic() cutoff
    (default 720s / 12 min, leaving a buffer against the 15-min Lambda
    ceiling) that every category's retry loop checks before each attempt.
    Without this, a generous per-category retry budget creates a new risk:
    one bad category retrying for a long time could push the WHOLE
    invocation past Lambda's hard timeout, which kills the process
    outright and loses every category's work, including the 9 that
    already succeeded. The shared deadline makes categories fail
    gracefully into "error" status individually instead, so generate()
    still returns a real (if partial) file rather than nothing at all.
  - bedrock_client.py's retry Config was also revisited in the same pass
    (see that file's own changelog) -- total_max_attempts raised further
    and retries mode switched to "adaptive", both leaning on the newly
    available time budget rather than failing fast.

PRIOR CHANGE LOG (retained for context):
  - RETRIEVAL REWORK (fixes slow/unreliable generation on large, normal
    -- i.e. non-scanned -- manuals): fetch_all_manual_chunks() used to
    pull EVERY chunk ever ingested for the equipment node, uncapped, and
    paste that single blob into all ten prompts unconditionally ("no
    vector search -- full recall", per the old docstring). A small
    scanned manual (a few thousand words) stayed fast; a large,
    genuinely dense native-text manual (tens of thousands of words, or
    an equipment node with more than one manual attached, since the old
    query wasn't even scoped to a single doc_id) turned into a huge
    prompt on all ten concurrent Bedrock calls -- slower, more likely to
    hit a timeout somewhere in the chain, and worse extraction quality
    from a small model (Haiku) reasoning over a much larger, noisier
    context. Replaced with fetch_relevant_manual_chunks(): a real
    pgvector similarity search, one per PM type, against a short
    type-specific query (see PM_QUERY_TEXT), capped at
    RETRIEVAL_TOP_K chunks. This uses the SAME embeddings already
    generated and stored per chunk at ingest time (see save_chunks() in
    ingest_document.py) -- they were always intended for retrieval, just
    not previously queried by vector similarity here. For a manual small
    enough that its total chunk count is under RETRIEVAL_TOP_K, this
    returns everything anyway, so small manuals (including the scanned
    ones going through Textract) see no behavior change at all.
  - Bedrock client now has an explicit timeout/retry Config
    (BEDROCK_CLIENT_CONFIG) instead of relying on botocore's undocumented
    default. Previously, a slow invoke_model call (more likely with the
    old unbounded context) would silently hit that default and get
    swallowed by call_claude_for_pm_type's broad except-clause, coming
    back as an empty PM type with no visible error -- indistinguishable
    from "Claude found nothing in this manual." The explicit
    read_timeout/connect_timeout below are sized for the now much
    smaller, bounded prompts this revision produces.
  - Image matching now tries an exact match on the manual's own part
    number (material_number) before falling back to the fuzzy
    keyword-in-page-text match on the component description. This is
    strictly additive: if material_number is blank or doesn't match
    anything, behavior is identical to before. See match_manual_image().
  - The material_number field instruction in build_pm_prompt was
    broadened from "SAP material number if mentioned" to also capture
    a manual's own parts-list number when the manual provides one --
    it was previously easy for Claude to leave this blank on manuals
    that don't use SAP terminology even when they clearly list part
    numbers (e.g. "341047").
  - fetch_all_manual_chunks / fetch_relevant_manual_chunks both handle
    the 4-segment content prefix ("equipment_id | doc_id | bom_items |
    text") written by the updated ingest_document.py, in addition to the
    older 2- and 3-segment formats, so this keeps working against both
    old and newly-ingested chunks without a backfill.
  - Image matching disabled via IMAGE_MATCHING_ENABLED (see note above).
  - Added generate_with_filename(), a thin wrapper around generate()
    that also returns a real filename instead of leaving that as a
    separate step for the caller to remember. generate() itself is
    unchanged in its return type (still bytes only) so nothing already
    calling it directly breaks.

NOTE ON WHERE THE "TIMEOUT SOMETIMES" WAS ACTUALLY COMING FROM: generation
already runs off a queue once the job is submitted -- API Gateway's 29s
ceiling only bounds the initial submit call, not this function, so it was
never the source of the reported timeouts. The real exposure was inside
this module itself: call_claude_for_pm_type()'s ten concurrent
invoke_model calls had no explicit timeout, so they inherited botocore's
undocumented default, and a slow call (more likely with the old unbounded,
full-recall prompt) would eventually hit that default and get swallowed by
the broad except-clause -- coming back as an empty PM type with no visible
error, indistinguishable from "Claude found nothing in this manual." That
still counts as a timeout from the caller's point of view even with no
API Gateway involved. The retrieval rework above shrinks and bounds every
prompt, and BEDROCK_CLIENT_CONFIG below makes the timeout explicit and
sized for that smaller prompt, so a genuinely slow call now fails fast and
visibly (logged, retried by botocore) instead of silently stalling out to
an empty result. Whatever timeout wraps the worker Lambda/consumer that
runs this job (its own function timeout, SQS visibility timeout, etc.)
should stay comfortably above BEDROCK_CLIENT_CONFIG's read_timeout * the
retry count, so a legitimately slow Bedrock call doesn't get killed by the
outer timeout before botocore's own retry/timeout logic has a chance to
resolve it. With total_max_attempts now at 6 (up from 2 retries), that
budget is larger than before -- worst case is roughly
read_timeout(90s) * 6 = 540s if every attempt were to hit the full read
timeout (unlikely in practice, since most retries fire fast on an
immediate ThrottlingException rather than waiting out the full timeout,
but budget the outer timeout as if it could happen).
"""

import io
import logging
import asyncio
import re
import time
from typing import Optional

import boto3
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.utils.db import get_db_connection
from app.services.embeddings import get_embedding
from app.services.bedrock_client import call_claude

log = logging.getLogger(__name__)

WORKING_PRINCIPLE = ("WP", "Working Principle")

# Working Principle is placed first: it's the foundational "how this
# equipment operates" narrative that the PM tasks below build on. It runs
# through the same async gather as the PM types but uses its own prompt
# (see build_working_principle_prompt) since it isn't a maintenance task
# extraction -- there's no frequency, work_needed, or failure mode in the
# same sense as PM1-PM9.
PM_TYPES = [
    WORKING_PRINCIPLE,
    ("PM1", "Inspection"),
    ("PM2", "Lubrication"),
    ("PM3", "Calibration"),
    ("PM4", "Replacements"),
    ("PM5", "Overhaul"),
    ("PM6", "Condition Monitoring"),
    ("PM7", "Cleaning"),
    ("PM8", "Safety Inspection"),
    ("PM9", "Software Back-up"),
]

# Short, type-specific queries used to retrieve the manual chunks most
# relevant to each PM type (see fetch_relevant_manual_chunks). These are
# deliberately written as a bag of concrete keywords/synonyms rather than
# a natural-language question -- they're only ever used to produce an
# embedding for similarity search, not read by a model, so density of
# relevant vocabulary matters more than grammar.
PM_QUERY_TEXT = {
    "WP":  "working principle theory of operation how the machine physically functions mechanism stages",
    "PM1": "inspection checking visual physical condition leaks tightness fluid level wear",
    "PM2": "lubrication greasing oil change lubricant grease coolant top up nipple",
    "PM3": "calibration adjustment pressure regulator governor engine speed torque specification setting",
    "PM4": "replacement scheduled part filter element belt fluid fastener consumable change interval",
    "PM5": "overhaul rebuild teardown reconditioning major internal assembly bearings rotors gears",
    "PM6": "condition monitoring measurement gauge sensor instrument threshold vibration temperature oil analysis",
    "PM7": "cleaning removing dirt grease debris buildup housing cooler core exterior wipe",
    "PM8": "safety inspection guard safety valve emergency stop decal label fastener panel",
    "PM9": "software backup firmware configuration PLC controller electronic control module restore",
}

# Model selection AND the Bedrock client's retry/timeout Config now live in
# app.services.bedrock_client (call_claude()), shared with /chat,
# ingest_document.py, and generate_job_aid.py, instead of being duplicated
# here. See that module's 2026-08-28 changelog entry for why (short
# version: this file used to build its own boto3 client with no Config at
# all, which meant every other caller of the shared client had the same
# undocumented-default-timeout exposure this incident was about, just
# unfixed). BEDROCK_REGION is kept here only because presign_manual_image_url()
# below needs a region for its own S3 client -- unrelated to Bedrock's region.
BEDROCK_REGION  = os.environ.get("AWS_REGION", "ca-central-1")

# Raised from 4096. Categories that legitimately produce 20+ multi-step
# tasks with numbered multi-line instructions (PM1, PM8 have both done this
# in production) can plausibly need more headroom than the old cap gave
# them. stop_reason is now logged per call inside call_claude() (see
# app.services.bedrock_client) so a truncated response is visible going
# forward instead of silently corrupting the parsed output.
MAX_TOKENS      = 8192

# Caps how many of THIS JOB's 10 category calls hit Bedrock at the same
# instant, instead of firing all 10 simultaneously via asyncio.gather.
# Lowered from 5 to 3: this Lambda has up to 15 minutes to work with and
# generation isn't user-facing/latency-sensitive, so there's no reason to
# maximize parallelism -- a smaller burst per job is strictly gentler on
# the shared Bedrock quota (see the multi-tenant discussion below) for a
# modest increase in this job's own wall-clock time, which is a trade this
# workload can easily afford. This only smooths the burst a single job
# creates by itself -- it does NOT coordinate across separate Lambda
# invocations or tenants. Bedrock's on-demand quota is account+region+
# model level, not per-tenant, so the real fix for one tenant's job
# throttling another tenant's job is an infra-level concurrency gate
# outside this file (a Lambda reserved-concurrency cap, or a shared token
# bucket in DynamoDB/ElastiCache) plus Bedrock cross-region inference to
# raise the effective ceiling. Treat this semaphore as a partial
# mitigation, not the fix for cross-tenant throttling.
BEDROCK_CALL_CONCURRENCY = int(os.environ.get("PM_BEDROCK_CALL_CONCURRENCY", 3))
_bedrock_semaphore = asyncio.Semaphore(BEDROCK_CALL_CONCURRENCY)

# How long, in seconds, generate() is willing to let its 10 category calls
# (including their retries -- see PM_CATEGORY_MAX_ATTEMPTS below) keep
# trying before giving up on whatever hasn't finished and returning a
# partial-but-real result. Set well under the Lambda function's actual
# timeout (per this deployment, up to 15 minutes = 900s) so a slow or
# repeatedly-throttled category produces a clean, logged "error" status on
# JUST that category instead of the WHOLE invocation being hard-killed by
# Lambda at 900s -- which would lose every category's work, including the
# ones that already succeeded. Defaults to 720s (12 min), leaving a
# 3-minute buffer for retrieval, Excel building, and whatever the caller
# does with the returned bytes afterward. Tune via env var if your actual
# Lambda timeout differs from 15 minutes -- this should always be
# comfortably below it, never equal to it.
GENERATION_DEADLINE_SECONDS = int(os.environ.get("PM_GENERATION_DEADLINE_SECONDS", 720))

# Application-level retries of a WHOLE category call (retrieval is NOT
# repeated on retry -- see call_claude_for_pm_type). Layered ON TOP OF
# call_claude()'s own botocore-level retries inside a single invoke_model
# call (network errors, throttling -- see bedrock_client.py's
# total_max_attempts). This is what actually uses the newly-available time
# budget: previously a category that got a usable HTTP response but an
# unusable one (e.g. no tool_use block) had zero retries and was
# permanently lost. 3 attempts, with linear backoff between them
# (PM_CATEGORY_RETRY_BACKOFF_SECONDS * attempt number, capped by whatever
# time is actually left against GENERATION_DEADLINE_SECONDS).
PM_CATEGORY_MAX_ATTEMPTS = int(os.environ.get("PM_CATEGORY_MAX_ATTEMPTS", 3))
PM_CATEGORY_RETRY_BACKOFF_SECONDS = 15

# How many of the most relevant chunks (by embedding similarity) to pull
# per PM type. ~40 chunks at ~500 words/chunk is roughly 20k words --
# generous headroom for a single maintenance category from most manuals,
# while still bounding worst-case prompt size on very large documents.
# For a manual with fewer than this many chunks total (true for every
# manual tested so far, including the scanned ones going through
# Textract), this simply returns everything -- no behavior change for
# the currently-working case.
RETRIEVAL_TOP_K = 40

# ── Structured output tool definition ──────────────────────────────────────
#
# Every WP/PM extraction call is forced (via tool_choice) to call this one
# tool instead of being asked to "return ONLY a valid JSON array" as free
# text. Anthropic parses and validates the arguments against this schema
# server-side; what comes back in the response is an ALREADY-PARSED object
# at content[*]["input"], not a string that this code has to bracket-find
# and json.loads() itself. This is what eliminates the WP "returned prose
# instead of JSON" failures and the PM8 json.JSONDecodeError -- both were
# symptoms of parsing model-generated text, and this removes that step
# entirely. Works on every Claude model on Bedrock (not just newer ones),
# so it's independent of whichever model app.services.bedrock_client is
# configured to use.
#
# The field descriptions here intentionally stay short -- the detailed
# per-field guidance (what counts as PM8 vs PM4, how to format the
# instruction field's numbered steps, etc.) lives in the prompt text in
# build_pm_prompt/build_working_principle_prompt, same as before. This
# schema only constrains shape and types; the prompt still teaches content.
EXTRACTION_TOOL = {
    "name": "record_extracted_steps",
    "description": (
        "Record the list of steps/tasks extracted from the equipment "
        "manual for this category. Call this even if no steps were found "
        "-- pass an empty tasks list rather than omitting the call."
    ),
    "input_schema": {
        "type": "object",
        "properties": {
            "tasks": {
                "type": "array",
                "description": "One entry per extracted step or task, in sequence order.",
                "items": {
                    "type": "object",
                    "properties": {
                        "operation": {
                            "type": "string",
                            "description": "Sequential step number, e.g. 'Operation_010', 'Operation_020'.",
                        },
                        "task_list_description": {"type": "string"},
                        "frequency": {"type": "string"},
                        "hrs": {
                            "type": ["number", "string"],
                            "description": "Decimal hours, or blank string if not specified.",
                        },
                        "work_needed": {"type": "integer"},
                        "system_condition": {"type": "integer"},
                        "material_number": {"type": "string"},
                        "component": {"type": "string"},
                        "instruction": {
                            "type": "string",
                            "description": "Numbered steps, each on its own line, separated by a blank line.",
                        },
                        "failure_modes": {"type": "string"},
                    },
                    "required": [
                        "operation", "task_list_description", "frequency", "hrs",
                        "work_needed", "system_condition", "material_number",
                        "component", "instruction", "failure_modes",
                    ],
                },
            },
        },
        "required": ["tasks"],
    },
}

COLUMNS = [
    "Operation",
    "Task List Description",
    "Frequency",
    "Hrs",
    "Work Needed",
    "System Condition",
    "Material Number",
    "Component",
    "Long Text (Instruction)",
    "Failure Modes",
    "Image",
]

HEADER_FILL   = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
SUBHEAD_FILL  = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")
SUBHEAD_FONT  = Font(bold=True, name="Arial", size=10)
DATA_FONT     = Font(name="Arial", size=10)
WRAP_ALIGN    = Alignment(wrap_text=True, vertical="top")

# Row/fill used to flag a category that errored during generation (as
# opposed to one that's legitimately empty because the manual doesn't
# cover it) -- see build_excel. Previously both looked identical: a
# header row followed by no data rows.
ERROR_FILL    = PatternFill("solid", start_color="C0392B", end_color="C0392B")
ERROR_FONT    = Font(bold=True, color="FFFFFF", name="Arial", size=10)

# Approx characters that fit on one wrapped line within the instruction
# column width (col width 60 -> roughly 60-65 chars/line at Arial 10).
INSTRUCTION_COL_CHARS_PER_LINE = 60
MIN_ROW_HEIGHT = 40
LINE_HEIGHT = 14  # points per wrapped line, Arial 10


# ── Knowledge retrieval ───────────────────────────────────────────────────────

_KNOWN_PREFIX_KEYS = ("equipment_id:", "doc_id:", "bom_items:")


def _strip_chunk_prefix(content: str) -> str:
    """
    Strip the leading "key:value | key:value | ..." metadata segments
    ingest_document.py writes into each chunk's content column, returning
    just the chunk's own text. Handles all three prefix shapes that exist
    across old and newly-ingested rows, oldest to newest:
      1. "equipment_id:{id} | {text}"                                (2 segments)
      2. "equipment_id:{id} | doc_id:{uuid} | {text}"                (3 segments)
      3. "equipment_id:{id} | doc_id:{uuid} | bom_items:... | {text}" (4 segments, current)
    Splitting on " | " with no maxsplit and consuming only the leading
    segments that match a known "key:" prefix keeps this working against
    whichever format a given chunk happens to be in, without needing to
    touch previously-ingested rows. Rejoins the remainder with " | " in
    case the chunk's own text legitimately contains that substring.
    """
    parts = content.split(" | ")
    split_at = 0
    for part in parts:
        if part.startswith(_KNOWN_PREFIX_KEYS):
            split_at += 1
        else:
            break
    if split_at == 0:
        return content  # no recognized prefix at all
    return " | ".join(parts[split_at:])


def _has_any_manual_chunks(equipment_id: str, company_id: str) -> bool:
    """
    Cheap existence check used to fail fast with a clear error before
    kicking off ten parallel embedding + retrieval + Bedrock calls for an
    equipment node that has no ingested manual at all.
    """
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT 1
                FROM knowledge_embeddings
                WHERE source_type = 'manual'
                AND company_id = %s::uuid
                AND content LIKE %s
                LIMIT 1
            """, (company_id, f"equipment_id:{equipment_id}%"))
            return cur.fetchone() is not None
    finally:
        conn.close()


def fetch_all_manual_chunks(equipment_id: str, company_id: str) -> str:
    """
    Fetch ALL manual chunks for this equipment node and concatenate them
    into a single text block. Kept for any direct/internal callers that
    want true full recall (e.g. a debugging tool, or a future feature
    that isn't per-PM-type) -- generate() itself no longer calls this;
    see fetch_relevant_manual_chunks() for what it uses instead, and the
    CHANGE LOG above for why.
    """
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT content
                FROM knowledge_embeddings
                WHERE source_type = 'manual'
                AND company_id = %s::uuid
                AND content LIKE %s
                ORDER BY created_at
            """, (company_id, f"equipment_id:{equipment_id}%"))
            rows = cur.fetchall()
    finally:
        conn.close()

    if not rows:
        raise ValueError(
            f"No manual chunks found for equipment {equipment_id}. "
            "Please upload and ingest a document for this node first."
        )

    clean_chunks = [_strip_chunk_prefix(r["content"]) for r in rows]
    full_text = "\n\n".join(clean_chunks)
    log.info(f"[company={company_id}] Fetched {len(clean_chunks)} chunks ({len(full_text.split())} words) for equipment {equipment_id}")
    return full_text


def fetch_relevant_manual_chunks(equipment_id: str, company_id: str, query_text: str, top_k: int = RETRIEVAL_TOP_K) -> str:
    """
    Vector-similarity retrieval of the top_k manual chunks most relevant
    to query_text, instead of dumping every chunk ever ingested for the
    equipment node into the prompt unconditionally. Uses the SAME
    embeddings already generated and stored per chunk at ingest time
    (see save_chunks() in ingest_document.py) -- this was always their
    intended purpose, just not previously queried here.

    Uses pgvector's cosine-distance operator (<=>), the standard choice
    for most text-embedding models. If the embeddings this deployment
    generates are tuned for a different distance metric, swap the
    operator below (<-> for Euclidean, <#> for inner product) to match.

    Falls back to an empty string (not an error) if this equipment has
    chunks but none happen to match well -- callers already handle an
    empty manual_text gracefully (Claude just returns an empty step list
    for that PM type, same as "the manual doesn't cover this category").
    The one-time upfront existence check in generate() is what catches
    "no manual ingested at all" -- this function assumes that's already
    been verified.
    """
    query_embedding = get_embedding(query_text)
    emb_str = "[" + ",".join(map(str, query_embedding)) + "]"

    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT content
                FROM knowledge_embeddings
                WHERE source_type = 'manual'
                AND company_id = %s::uuid
                AND content LIKE %s
                ORDER BY embedding <=> %s::vector
                LIMIT %s
            """, (company_id, f"equipment_id:{equipment_id}%", emb_str, top_k))
            rows = cur.fetchall()
    finally:
        conn.close()

    clean_chunks = [_strip_chunk_prefix(r["content"]) for r in rows]
    full_text = "\n\n".join(clean_chunks)
    log.info(
        f"[company={company_id}] Retrieved {len(clean_chunks)}/{top_k} chunks ({len(full_text.split())} words) "
        f"for equipment {equipment_id}, query={query_text[:60]!r}"
    )
    return full_text


# ── Claude call ───────────────────────────────────────────────────────────────

def build_working_principle_prompt(manual_text: str) -> str:
    """
    Working Principle answers "how does this machine work?" -- the
    engineering logic of how the equipment physically achieves its
    function, NOT how an operator runs it and NOT maintenance tasks.

    The example below is intentionally abstract/generic (no real
    equipment domain, no specific component names) rather than a
    concrete worked example (e.g. an air compressor). An earlier
    version used a real compressor example, and the model latched
    onto that specific vocabulary (screw airend, oil injection,
    receiver-separator) and started projecting compressor terminology
    onto completely unrelated equipment. The example here only
    demonstrates the reasoning style expected -- every actual term must
    come from the manual provided, never from this example.

    NOTE: this used to end with a "Return ONLY a valid JSON array" +
    literal JSON example block. That's no longer needed -- the caller
    forces this response through the record_extracted_steps tool (see
    EXTRACTION_TOOL), which constrains the shape server-side. The
    field-by-field guidance below still matters (it's what teaches the
    model what belongs in each field); only the now-redundant formatting
    instructions were removed.
    """
    return f"""You are a maintenance engineering expert. You have been given an equipment manual below.

Your task is to extract the WORKING PRINCIPLE of THIS SPECIFIC MACHINE -- the engineering explanation of HOW IT PHYSICALLY ACHIEVES ITS FUNCTION. This answers the question "how does this machine work?", not "how does an operator run it?"

Do NOT extract:
- Operator procedures: button presses, switch positions, start/stop sequences, towing/setup instructions.
- Maintenance tasks: inspection, lubrication, replacement, cleaning, calibration.

DO extract the underlying mechanism: how the machine's core function is physically carried out, stage by stage, as material/air/fluid/energy/signal moves through the system and components interact to produce the intended output. The specific stages, components, and terminology MUST come entirely from the manual below -- do not import terminology, component names, or mechanisms from any other type of equipment. If this machine is not a compressor, do not mention compression, airends, or oil separation; if it's not a pump, do not mention impellers; describe THIS machine using only what the manual actually says about it.

IMPORTANT: manuals rarely have a section explicitly titled "theory of operation" or "working principle." This kind of engineering explanation is usually scattered as incidental description INSIDE maintenance, general data, or specification sections -- not confined to a section literally called "Operation." Read the whole manual for sentences that explain WHY or HOW something physically happens, even if that sentence sits inside a section about draining or replacing a part. Extract the engineering logic from wherever it appears in THIS manual; do not limit yourself to a single section, and do not substitute reasoning from a different type of machine when this manual's explanation is thin -- if the manual genuinely doesn't explain a stage, leave it out rather than inventing it.

For each functional stage, extract:
- operation: sequential step number as "Operation_010", "Operation_020", "Operation_030" etc (increment by 10), in the order the function is physically carried out
- task_list_description: the functional stage following the pattern "Function - Mechanism", using the manual's own terms for the function and the component that carries it out
- frequency: leave blank (not applicable)
- hrs: leave blank (not applicable)
- work_needed: 0 (this describes function, not maintenance work)
- system_condition: 1 (this describes the machine's normal running function)
- material_number: leave blank
- component: the specific component or subsystem named in the manual that carries out this function
- instruction: a clear, detailed engineering explanation of what physically happens at this stage and why, using only terms and mechanisms the manual actually describes. Number each sub-point starting at 1. Put EACH numbered point on its own line, with a blank line between points -- use a literal "\\n\\n" (newline, newline) between point N and point N+1, never a space.
- failure_modes: leave blank (not applicable)

Call the record_extracted_steps tool with your findings. If the manual has no engineering description of how the machine functions anywhere, call it with an empty tasks list.

EQUIPMENT MANUAL:
{manual_text}"""


PM_TYPE_GUIDANCE = {
    "PM1": "Inspection means a routine visual or physical CHECK with no scheduled part replacement -- e.g. checking for leaks, checking fastener tightness, checking fluid level, visually inspecting hoses or tires. If the task's end action is replacing a part, it does NOT belong here -- that's PM4.",
    "PM2": "Lubrication means applying, topping up, or changing a lubricant, grease, or coolant itself (e.g. greasing a fitting, topping up engine oil, changing compressor oil). Do not include filter/element replacement here unless the task is specifically about the fluid, not the filter -- filter/element swaps belong in PM4.",
    "PM3": "Calibration means adjusting a device or setting to a specified reference value. This includes pressure regulator adjustment, governor or engine speed adjustment, sensor calibration, torque specifications, and any 'adjust to X value' instruction in the manual -- even if the manual's section heading doesn't use the word 'calibration'. If the manual has any section on adjusting pressure, speed, torque, or regulator settings, it belongs here.",
    "PM4": "Replacements means the routine, scheduled swap of a wearable part or consumable on a fixed interval (filters, elements, belts, fluids, fasteners). This is the ONLY category that should contain recurring element/filter/fluid replacement tasks -- do not also list the same task under PM2, PM5, or PM6.",
    "PM5": "Overhaul means major teardown, rebuild, or reconditioning of an assembly (e.g. rebuilding the airend internals, engine rebuild, replacing internal bearings/rotors/gears as part of a rebuild). Do NOT include routine scheduled part replacement here (that's PM4) or routine fluid changes (that's PM2). If the manual explicitly states major overhauls are outside its scope or should be referred to an authorized service department, return an EMPTY array for this category -- do not invent overhaul tasks by relabeling routine maintenance content.",
    "PM6": "Condition Monitoring means measuring or observing a parameter against a threshold using a gauge, sensor, or instrument, WITHOUT performing a replacement as part of the same task (e.g. monitoring discharge temperature, monitoring vibration, oil sampling/analysis, watching a diagnostic lamp). If a task's end action is replacing a part on a schedule, it belongs in PM4, not here.",
    "PM7": "Cleaning means removing dirt, grease, debris, or buildup from a component that stays installed (e.g. cleaning a cooler core exterior, wiping down a housing interior). Do not include tasks whose primary action is replacing a part.",
    "PM8": "Safety Inspection means checking safety-critical items specifically: guards, safety valves, safety decals/labels, emergency stops, and fasteners or panels tied to injury or noise-containment risk.",
    "PM9": "Software Back-up means backing up or restoring configuration, firmware, or software on a PLC, controller, or electronic control module. If the manual describes no software, controller, or firmware, return an EMPTY array -- do not invent a software task.",
}


def build_pm_prompt(pm_code: str, pm_name: str, manual_text: str) -> str:
    """
    NOTE: this used to end with a "Return ONLY a valid JSON array" +
    literal JSON example block, same as build_working_principle_prompt
    above. Removed for the same reason -- tool_choice now enforces the
    output shape, so that instruction block was redundant (and one of
    the sources of drift: the model would sometimes narrate instead of
    strictly following it). The category guidance below is unchanged.
    """
    category_guidance = PM_TYPE_GUIDANCE.get(pm_code, "")
    return f"""You are a maintenance engineering expert. You have been given an equipment manual below.

Your task is to extract ALL maintenance tasks that fall under the category: {pm_code} - {pm_name}

CATEGORY DEFINITION FOR {pm_code} - {pm_name}: {category_guidance}

If a task genuinely fits more than one category, extract it under the SINGLE most specific category above and skip it in the others -- do not extract the same task into multiple PM types.

For each task you find, extract the following fields:
- operation: sequential step number as "Operation_010", "Operation_020", "Operation_030" etc (increment by 10)
- task_list_description: the step title following the component hierarchy pattern "Assembly - Subassembly - Component x[quantity]". Preserve quantities (x1, x2, x4 etc) as they indicate how many of that component exist.
- frequency: how often this task should be done (e.g. "2W" for 2 weekly, "1M" for monthly, "1Y" for yearly). Leave blank if not specified.
- hrs: estimated TECHNICIAN TIME to perform this specific task, as a decimal number of hours (e.g. 0.1, 0.5, 1.0). This is NOT the maintenance interval -- if the manual says "every 1000 hours," that 1000 belongs in frequency, never here. Leave blank if no time estimate is given.
- work_needed: 1 if active work is required, 0 if observation only. Default to 1.
- system_condition: 0 for machine stopped, 1 for machine running. Default to 0.
- material_number: the manual's OWN part/component number for this item, if the manual provides one -- this includes a formal SAP material number, but just as importantly it includes any parts-list, drawing, or catalog number the manual itself prints next to this component (e.g. "341047", "992-03377", "8-110-626-107"). Copy it exactly as printed, including hyphens/periods. Leave blank only if the manual truly gives no such number for this component.
- component: the specific component name only (without the assembly hierarchy), e.g. "Bearings x8"
- instruction: step-by-step work instruction a technician can follow. Number each step starting at 1. Put EACH numbered step on its own line, with a blank line between steps -- use a literal "\\n\\n" (newline, newline) between step N and step N+1, never a space. Be specific.
- failure_modes: comma-separated list of failure modes this task prevents. Leave blank if not specified.

Call the record_extracted_steps tool with your findings. If no tasks of type {pm_name} are found in the manual, call it with an empty tasks list.

EQUIPMENT MANUAL:
{manual_text}"""


# ── Component image lookup ───────────────────────────────────────────────────
#
# Images come ONLY from this equipment's own ingested manual
# (equipment_manual_images table, populated at ingestion time by
# ingest_document.py's PDF image extraction). No external image search
# (Openverse or otherwise) -- if the manual has no relevant image, the
# cell is left blank for the reviewer to fill, rather than substituting
# a generic stock photo that isn't actually this equipment's part.
#
# Matching is now two-tier:
#   1. Exact match on material_number (the manual's own part number, e.g.
#      "341047") against a drawing page's extracted text. This is a much
#      stronger, lower-false-positive signal than a keyword match, since
#      part numbers are specific tokens rather than descriptive words that
#      might appear on several unrelated pages ("bearing" turning up on
#      six different drawings). It only fires when Claude actually
#      populated material_number for that step AND that exact string
#      appears on some page's extracted text -- otherwise it's a silent
#      no-op and behavior falls through to tier 2, unchanged from before.
#   2. Fallback: keyword-in-page-text match on the component description,
#      same coarse approach as before.
#
# This is a fully local operation: the S3 key is already in the DB, and
# generating a presigned URL is a local signing call, not a network
# request. There's no timeout, retry, or concurrency-limiting logic
# needed here because there's nothing that can hang or rate-limit.
#
# Image lookup happens INLINE inside call_claude_for_pm_type(), right
# after each PM/WP type's steps come back from Claude. WP is included:
# its "component" field names real physical parts (e.g. "Airend",
# "Pressure Regulator"), same as PM1-PM9's.
#
# Known limitation, by design: match quality depends on the ingested
# manual actually having a relevant image AND that image's page text
# mentioning the component (by part number or by name). A manual with
# no images, or a components section whose images sit on pages that
# don't share the component's wording or number, will come back blank --
# there's no fallback beyond tier 2.

# Master switch for image matching, mirroring ingest_document.py's
# IMAGE_EXTRACTION_ENABLED. Off for now because ingest no longer writes
# to equipment_manual_images -- querying it here would just be a wasted
# round-trip against a table that's always empty. This only gates the
# fetch in generate() below; match_manual_image/resolve_component_image_url/
# attach_images_to_steps are untouched and already degrade correctly to
# all-blank image_urls when handed an empty manual_images list, so
# there's nothing else to change when this flips back to True (besides
# also re-enabling IMAGE_EXTRACTION_ENABLED in ingest_document.py so
# there's actually something to fetch).
IMAGE_MATCHING_ENABLED = False

S3_BUCKET = os.environ.get("S3_BUCKET", "squaremethods")
MANUAL_IMAGE_URL_EXPIRY_SECONDS = 604800  # 7 days -- the SigV4 max; images stay private, not permanently public

_QUANTITY_SUFFIX_RE = re.compile(r"\s*[xX]\d+\s*$")

# Guard against spurious substring matches on short/generic material
# numbers (e.g. a stray "24" matching all over a page of tables). Real
# manual part numbers in the manuals we've seen are consistently longer
# than this, so it's a cheap, conservative filter rather than a tuned one.
MIN_PART_NUMBER_MATCH_LENGTH = 4


def normalize_component_for_image_search(component: str) -> Optional[str]:
    """
    Turn a Component field value into a bare search keyword.
    "Main Drive assembly - Bearings x4"  -> "bearings"
    "Compressor Oil Filter Element"      -> "compressor oil filter element"
    "Bearings x8"                       -> "bearings"
    Returns None for empty/unusable input so callers can skip the lookup.
    """
    if not component:
        return None

    # Strip assembly/subassembly hierarchy -- keep only the last segment
    # after " - ", since that's the actual component, not its location.
    term = component.split(" - ")[-1]

    # Strip trailing quantity markers like "x4", "X12"
    term = _QUANTITY_SUFFIX_RE.sub("", term)

    term = term.strip()
    return term.lower() if term else None


def fetch_manual_images_for_equipment(equipment_id: str, company_id: str) -> list:
    """
    Pulls every extracted manual image row for this equipment ONCE, up
    front, so matching against many components doesn't mean many DB
    round-trips. Returns a list of {s3_key, context_text, page_number}
    dicts. Empty list if the equipment has no ingested images (e.g. a
    DOCX manual, or a PDF with no images that cleared the size filter).
    """
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT s3_key, context_text, page_number
                FROM equipment_manual_images
                WHERE equipment_id = %s::uuid
                AND company_id = %s::uuid
            """, (equipment_id, company_id))
            rows = cur.fetchall()
    finally:
        conn.close()

    log.info(f"[company={company_id}] Loaded {len(rows)} manual images for equipment {equipment_id}")
    return [
        {"s3_key": r["s3_key"], "context_text": (r["context_text"] or "").lower(), "page_number": r["page_number"]}
        for r in rows
    ]


def match_manual_image(component_keyword: str, material_number: str, manual_images: list) -> tuple:
    """
    Two-tier match against this equipment's own ingested manual images.
    Returns (s3_key, match_method) where match_method is "part_number",
    "keyword", or None if nothing matched. Returning the method alongside
    the key is just for the summary logging in call_claude_for_pm_type --
    callers that don't care can ignore it.
    """
    if not manual_images:
        return None, None

    part_no = (material_number or "").strip().lower()
    if len(part_no) >= MIN_PART_NUMBER_MATCH_LENGTH:
        for img in manual_images:
            if part_no in img["context_text"]:
                return img["s3_key"], "part_number"

    if component_keyword:
        for img in manual_images:
            if component_keyword in img["context_text"]:
                return img["s3_key"], "keyword"

    return None, None


def presign_manual_image_url(s3_key: str) -> Optional[str]:
    """
    Generate a presigned GET URL for a manual image. Images are stored
    privately (see ingest_document.py) since manuals may be confidential
    OEM documents -- this is a local signing operation, not a network
    call, so it can't hang or fail on connectivity.
    """
    try:
        s3 = boto3.client("s3", region_name=BEDROCK_REGION)
        return s3.generate_presigned_url(
            "get_object",
            Params={"Bucket": S3_BUCKET, "Key": s3_key},
            ExpiresIn=MANUAL_IMAGE_URL_EXPIRY_SECONDS,
        )
    except Exception as e:
        log.warning(f"Failed to presign manual image URL for key {s3_key}: {type(e).__name__}: {e}")
        return None


def resolve_component_image_url(component_keyword: str, material_number: str, manual_images: list) -> tuple:
    """
    Fully synchronous and local -- no network call, so nothing here can
    hang or need a timeout. Returns (url, match_method); url is "" if
    the equipment's manual has no matching image, in which case the
    Excel cell is simply left blank.
    """
    manual_key, method = match_manual_image(component_keyword, material_number, manual_images)
    if not manual_key:
        return "", None
    return presign_manual_image_url(manual_key) or "", method


def attach_images_to_steps(steps: list, manual_images: list) -> dict:
    """
    Mutates each step dict in place with an "image_url" key, matched
    against this equipment's own ingested manual images. Purely
    synchronous -- no async/await needed since there's no network call
    involved anywhere in this path.

    Returns a small {"part_number": n, "keyword": n} tally for logging.
    """
    tally = {"part_number": 0, "keyword": 0}
    for step in steps:
        keyword = normalize_component_for_image_search(step.get("component", ""))
        material_number = step.get("material_number", "")
        url, method = resolve_component_image_url(keyword or "", material_number, manual_images)
        step["image_url"] = url
        if method:
            tally[method] += 1
    return tally


async def _attempt_bedrock_call(prompt: str, company_id: str, pm_code: str) -> tuple[Optional[list], Optional[str], dict]:
    """
    One attempt at the Bedrock call + response parsing for a single
    category. Split out of call_claude_for_pm_type so the retry loop
    below can call this repeatedly without re-running retrieval (which
    is deterministic and not worth repeating) each time.

    Returns (steps_or_None, error_reason_or_None, raw_response_dict).
    steps is None (not []) specifically when this attempt failed --
    an empty list is a legitimate "Claude found nothing" result and
    should NOT be retried, only a None/error result should be.
    """
    loop = asyncio.get_event_loop()
    try:
        # Capped so this job's category calls don't all fire in the same
        # instant -- see BEDROCK_CALL_CONCURRENCY for what this does and
        # does not solve. call_claude() (app.services.bedrock_client) owns
        # the actual boto3 client, its retry/timeout Config, and the
        # raw-body/stop_reason/usage logging -- /chat and ingest_document.py
        # get the same hardening through that same shared call.
        async with _bedrock_semaphore:
            raw = await loop.run_in_executor(
                None,
                lambda: call_claude(
                    messages=[{"role": "user", "content": prompt}],
                    tools=[EXTRACTION_TOOL],
                    tool_choice={"type": "tool", "name": EXTRACTION_TOOL["name"]},
                    max_tokens=MAX_TOKENS,
                    temperature=0,
                    log_context=f"[company={company_id}] {pm_code}",
                ),
            )
    except Exception as invoke_err:
        log.error(f"[company={company_id}] {pm_code} CALL_CLAUDE FAILED: {type(invoke_err).__name__}: {invoke_err}")
        return None, f"{type(invoke_err).__name__}: {invoke_err}", {}

    # call_claude() already logged raw body length, stop_reason, and token
    # usage (with the log_context prefix above).
    stop_reason = raw.get("stop_reason")

    # Structured output via forced tool use: the model's arguments come
    # back as an ALREADY-PARSED object in a "tool_use" content block's
    # "input" field, not as text this code has to bracket-find and
    # json.loads() itself. This is what removes the whole class of bug
    # that hit WP (returned prose instead of JSON) and PM8 (malformed
    # JSON from an escaping mistake) -- there's no free-text parsing step
    # left for either failure mode to occur in.
    steps = None
    for block in raw.get("content", []):
        if block.get("type") == "tool_use" and block.get("name") == EXTRACTION_TOOL["name"]:
            steps = block.get("input", {}).get("tasks")
            break

    if steps is None:
        # tool_choice forces this tool, so its absence is a real anomaly
        # worth retrying rather than a quiet "empty" -- most likely a
        # response that got cut off (check stop_reason) before the tool
        # call was emitted.
        reason = f"no tool_use block in response, stop_reason={stop_reason!r}"
        log.error(
            f"[company={company_id}] {pm_code} {reason}. "
            f"Raw content preview: {str(raw.get('content'))[:300]}"
        )
        return None, reason, raw
    if not isinstance(steps, list):
        reason = f"tool_use input.tasks was not a list: {type(steps).__name__}"
        log.error(f"[company={company_id}] {pm_code} {reason}")
        return None, reason, raw

    return steps, None, raw


async def call_claude_for_pm_type(
    pm_code: str,
    pm_name: str,
    equipment_id: str,
    company_id: str,
    manual_images: list,
    deadline: float,
) -> tuple[str, str, list, str]:
    """
    Returns (pm_code, pm_name, steps, status) where status is one of:
      "ok"    -- steps were extracted normally
      "empty" -- Claude legitimately found nothing for this category
                 (a valid, expected outcome, not a failure) -- NOT retried
      "error" -- every attempt failed (retrieval, invoke_model, or a
                 response that never produced a usable tool_use block) --
                 previously indistinguishable from "empty" in the output,
                 and previously never retried at all.
    See build_excel for how "error" is surfaced in the spreadsheet.

    deadline: a time.monotonic() cutoff shared across all 10 categories
    in this job (see GENERATION_DEADLINE_SECONDS / generate()). Checked
    before every attempt, including the first, so that if this job is
    already running long by the time this particular category's task
    gets scheduled, it fails fast and gracefully into "error" rather than
    starting an attempt it has no realistic time budget to retry. This is
    what keeps one slow/throttled category from eating the time every
    other category needed too, and stops the whole Lambda invocation from
    being hard-killed at its timeout with nothing to show for it --
    see the module changelog for why this matters more now that
    individual attempts are allowed to take their time rather than fail
    fast.
    """
    log.debug(f"[company={company_id}] {pm_code} ENTERED call_claude_for_pm_type, retrieval starting now")
    loop = asyncio.get_event_loop()

    # Retrieval runs per-PM-type and concurrently with the other nine
    # (via run_in_executor, since get_embedding()/psycopg2 are blocking
    # calls) rather than once upfront for the whole document -- see the
    # CHANGE LOG at the top of this file for why full-recall was replaced
    # with per-type similarity search. Done ONCE here, not repeated across
    # retries below -- it's deterministic (same query, same embeddings,
    # same top-K), so re-running it would only burn time budget for no
    # benefit.
    query_text = PM_QUERY_TEXT.get(pm_code, pm_name)
    try:
        manual_text = await loop.run_in_executor(
            None, fetch_relevant_manual_chunks, equipment_id, company_id, query_text
        )
    except Exception as retrieval_err:
        log.error(f"[company={company_id}] {pm_code} RETRIEVAL FAILED: {type(retrieval_err).__name__}: {retrieval_err}")
        return pm_code, pm_name, [], "error"

    if pm_code == "WP":
        prompt = build_working_principle_prompt(manual_text)
    else:
        prompt = build_pm_prompt(pm_code, pm_name, manual_text)
    log.debug(f"[company={company_id}] {pm_code} PROMPT BUILT, length={len(prompt)}")
    log.info(f"[company={company_id}] {pm_code} PROMPT LENGTH (chars): {len(prompt)}")

    # Application-level retry loop, layered ON TOP OF call_claude()'s own
    # botocore-level retries (network errors, throttling within a single
    # invoke_model call -- see bedrock_client.py). This loop instead
    # catches "the call succeeded at the network level but the response
    # wasn't usable" -- a fresh sample from the model on the next attempt
    # often just avoids whatever made the first one unusable. Previously
    # there was no retry here at all: one bad response permanently lost
    # that category. Now that speed isn't the priority (up to 15 minutes
    # available in Lambda per the infra this runs on), this trades time
    # budget for a real second and third chance instead of giving up
    # after one shot.
    steps: list = []
    status = "error"
    last_reason = "not attempted"
    for attempt in range(1, PM_CATEGORY_MAX_ATTEMPTS + 1):
        remaining = deadline - time.monotonic()
        if remaining <= 0:
            log.error(
                f"[company={company_id}] {pm_code} out of time budget before attempt {attempt}/"
                f"{PM_CATEGORY_MAX_ATTEMPTS} (last failure: {last_reason}) -- giving up on this category"
            )
            break

        if attempt > 1:
            log.warning(
                f"[company={company_id}] {pm_code} attempt {attempt}/{PM_CATEGORY_MAX_ATTEMPTS} "
                f"(previous attempt failed: {last_reason})"
            )

        result_steps, error_reason, _raw = await _attempt_bedrock_call(prompt, company_id, pm_code)

        if error_reason is None:
            # result_steps is a real list here, possibly empty -- an
            # empty list is Claude legitimately finding nothing, which is
            # a valid outcome and NOT retried.
            steps = result_steps
            status = "ok" if steps else "empty"
            break

        last_reason = error_reason
        if attempt < PM_CATEGORY_MAX_ATTEMPTS:
            # Linear backoff between application-level retries (separate
            # from, and on top of, botocore's own backoff inside a single
            # call_claude() invocation). Capped against whatever time
            # budget is actually left so this never itself becomes the
            # reason the shared deadline gets blown.
            backoff = min(PM_CATEGORY_RETRY_BACKOFF_SECONDS * attempt, max(remaining - 5, 0))
            if backoff > 0:
                await asyncio.sleep(backoff)

    if status == "error":
        log.error(
            f"[company={company_id}] {pm_code} exhausted all attempts, last failure: {last_reason}"
        )

    log.info(f"[company={company_id}] {pm_code} ({pm_name}): {len(steps)} steps extracted, status={status}")

    # Resolve manual images for this type's steps now. Purely local (DB
    # already fetched, presigning is local signing) -- no network call, so
    # this can't hang. Runs for WP too -- its components (e.g. "Airend",
    # "Pressure Regulator") are genuine physical parts, same as PM1-PM9's.
    tally = attach_images_to_steps(steps, manual_images)
    matched = tally["part_number"] + tally["keyword"]
    if steps:
        log.info(
            f"[company={company_id}] {pm_code} images: {matched}/{len(steps)} steps matched "
            f"({tally['part_number']} by part number, {tally['keyword']} by keyword)"
        )

    return pm_code, pm_name, steps, status


# ── Excel builder ─────────────────────────────────────────────────────────────

def _estimate_row_height(instruction: str) -> int:
    """
    Instruction text contains numbered steps separated by blank lines
    (\\n\\n). Estimate wrapped line count so multi-step instructions
    don't get visually clipped at a fixed row height.
    """
    if not instruction:
        return MIN_ROW_HEIGHT

    line_count = 0
    for segment in instruction.split("\n"):
        if segment == "":
            line_count += 1  # blank line between steps
        else:
            # account for wrapping within a single step line
            line_count += max(1, -(-len(segment) // INSTRUCTION_COL_CHARS_PER_LINE))

    return max(MIN_ROW_HEIGHT, line_count * LINE_HEIGHT)


def build_excel(equipment_id: str, pm_results: list[tuple]) -> bytes:
    """
    Build the PM strategy Excel file from Claude's structured output.
    Format matches PM_strategy.xlsx exactly.

    pm_results: list of (pm_code, pm_name, steps_list, status) tuples in
    PM1-PM9 order. status is "ok" / "empty" / "error" (see
    call_claude_for_pm_type). All PM types are written -- empty ones show
    the header and column row with no data rows, ready for the reviewer
    to fill in manually. A category with status "error" gets a visibly
    flagged row instead of looking identical to a legitimately empty one
    -- previously there was no way to tell the two apart just by looking
    at the file.
    """
    wb   = Workbook()
    ws   = wb.active
    ws.title = "PM Strategy"

    col_widths = [15, 55, 12, 8, 13, 17, 16, 35, 60, 35, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    current_row = 1

    for pm_code, pm_name, steps, status in pm_results:

        # PM type header row (e.g. "PM2 - Lubrication")
        header_cell = ws.cell(row=current_row, column=1, value=f"{pm_code} - {pm_name}")
        header_cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        header_cell.fill = HEADER_FILL
        header_cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[current_row].height = 20
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=len(COLUMNS)
        )
        current_row += 1

        # Column headers
        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell           = ws.cell(row=current_row, column=col_idx, value=col_name)
            cell.font      = SUBHEAD_FONT
            cell.fill      = SUBHEAD_FILL
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        ws.row_dimensions[current_row].height = 30
        current_row += 1

        if status == "error":
            # Flagged distinctly from a legitimately empty category so the
            # reviewer (or the salesperson sending this to a customer)
            # doesn't mistake "generation failed for this category" for
            # "the manual just doesn't cover this".
            error_cell = ws.cell(
                row=current_row, column=1,
                value=f"GENERATION ERROR - {pm_code} failed to generate, retry this category. Do not treat as empty.",
            )
            error_cell.font = ERROR_FONT
            error_cell.fill = ERROR_FILL
            error_cell.alignment = Alignment(vertical="center")
            ws.row_dimensions[current_row].height = 20
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=len(COLUMNS)
            )
            current_row += 1

        # Data rows -- skipped naturally if steps is empty
        for step in steps:
            instruction = step.get("instruction", "")
            image_url = step.get("image_url", "")
            row_values = [
                step.get("operation", ""),
                step.get("task_list_description", ""),
                step.get("frequency", ""),
                step.get("hrs", ""),
                step.get("work_needed", ""),
                step.get("system_condition", ""),
                step.get("material_number", ""),
                step.get("component", ""),
                instruction,
                step.get("failure_modes", ""),
                image_url,  # manual reference URL if matched, blank otherwise -- reviewer edits/replaces as needed
            ]
            for col_idx, value in enumerate(row_values, 1):
                cell           = ws.cell(row=current_row, column=col_idx, value=value)
                cell.font      = DATA_FONT
                cell.alignment = WRAP_ALIGN
                if col_idx == len(COLUMNS) and image_url:
                    cell.hyperlink = image_url
                    cell.font = Font(name="Arial", size=10, color="0563C1", underline="single")
            ws.row_dimensions[current_row].height = _estimate_row_height(instruction)
            current_row += 1

        # Spacer row between PM type blocks
        current_row += 1

    # Freeze the top row of the first block
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def fetch_equipment_info(equipment_id: str, company_id: str) -> dict:
    """
    Pull the equipment's display name and reference code for use in the
    output filename. Matches the `equipment` table schema: `name` is the
    human-readable label, `reference_code` is the unique per-company
    identifier -- both go in the filename so two units with the same
    name (different reference_code) don't produce identical filenames.
    Respects the soft-delete pattern (deleted_at IS NULL) used elsewhere
    in the equipment import system.
    """
    conn = get_db_connection()
    try:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT name, reference_code
                FROM equipment
                WHERE id = %s::uuid
                AND company_id = %s::uuid
                AND deleted_at IS NULL
            """, (equipment_id, company_id))
            row = cur.fetchone()
    finally:
        conn.close()

    if not row:
        log.warning(f"[company={company_id}] No equipment record found for {equipment_id}, falling back to equipment_id in filename")
        return {"name": "", "reference_code": ""}

    return {"name": row.get("name") or "", "reference_code": row.get("reference_code") or ""}


def build_output_filename(equipment_id: str, equipment_info: dict) -> str:
    """
    Build a human-readable filename like:
      PM_Strategy_P185WJD_Compressor_1_A102_2026-07-26.xlsx
    (name + reference_code). Falls back to the equipment_id if no
    record is found, so the file is still uniquely identifiable rather
    than failing.
    """
    from datetime import date

    label = " ".join(part for part in [equipment_info.get("name"), equipment_info.get("reference_code")] if part).strip()
    if not label:
        label = equipment_id

    # slugify: keep letters/numbers, collapse everything else to a single underscore
    slug = re.sub(r"[^A-Za-z0-9]+", "_", label).strip("_")
    slug = slug[:80]  # keep filenames reasonable on Windows/shared drives

    return f"PM_Strategy_{slug}_{date.today().isoformat()}.xlsx"


# ── Main entry point ──────────────────────────────────────────────────────────

async def generate(equipment_id: str, company_id: str) -> bytes:
    """
    Full pipeline called from the FastAPI endpoint.
    1. Fail fast if this equipment has no ingested manual at all
    2. Run ten parallel Claude calls (Working Principle + nine PM types),
       each retrieving its own most-relevant manual chunks (see
       fetch_relevant_manual_chunks) and resolving its own stock
       component images inline. Concurrency to Bedrock is capped at
       BEDROCK_CALL_CONCURRENCY (see that constant's comment for what
       this does and does not solve).
    3. Build and return the Excel file as bytes
       -- always returns a valid file even if all PM types are empty;
       any category that errored (as opposed to legitimately empty) is
       flagged visibly in the file rather than looking identical to an
       empty one -- see build_excel.

    Return type is unchanged (bytes only) so nothing already calling
    generate() directly breaks. For a real filename bundled with the
    bytes instead of a generic "download.xlsx", call
    generate_with_filename() instead -- see that function below.

    Example endpoint usage (preferred -- real filename included):
        excel_bytes, filename = await generate_with_filename(equipment_id, company_id)
        return Response(
            content=excel_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    """
    if not _has_any_manual_chunks(equipment_id, company_id):
        raise ValueError(
            f"No manual chunks found for equipment {equipment_id}. "
            "Please upload and ingest a document for this node first."
        )

    # Fetched once, up front -- this equipment's own manual images are
    # matched in-memory against every step's component keyword, rather
    # than one DB round-trip per keyword. No network client needed for
    # image lookup anymore: it's a local list match + local S3 presign.
    #
    # Skipped entirely while IMAGE_MATCHING_ENABLED is False: ingest no
    # longer writes to equipment_manual_images, so this table is always
    # empty right now and querying it would just be wasted latency. The
    # downstream matching code already handles an empty list correctly
    # (every step's image_url comes back ""), so this is the only line
    # that needs to change to turn image matching off end-to-end.
    manual_images = fetch_manual_images_for_equipment(equipment_id, company_id) if IMAGE_MATCHING_ENABLED else []

    # Shared across every category's retry loop -- see GENERATION_DEADLINE_SECONDS
    # and call_claude_for_pm_type's docstring for why this needs to be one
    # deadline all 10 tasks check against, not 10 independent budgets.
    deadline = time.monotonic() + GENERATION_DEADLINE_SECONDS

    tasks = [
        call_claude_for_pm_type(pm_code, pm_name, equipment_id, company_id, manual_images, deadline)
        for pm_code, pm_name in PM_TYPES
    ]
    results = await asyncio.gather(*tasks)

    pm_results = list(results)

    ok_count    = sum(1 for _, _, _, status in pm_results if status == "ok")
    empty_count = sum(1 for _, _, _, status in pm_results if status == "empty")
    error_count = sum(1 for _, _, _, status in pm_results if status == "error")
    log.info(
        f"[company={company_id}] Generation complete for equipment {equipment_id}. "
        f"{ok_count} ok / {empty_count} legitimately empty / {error_count} errored "
        f"(of {len(PM_TYPES)} PM types)."
    )
    if error_count:
        log.warning(
            f"[company={company_id}] {error_count} PM type(s) errored during generation for "
            f"equipment {equipment_id} -- these are flagged in the output Excel, but this job "
            f"should probably be retried or surfaced to the caller rather than treated as fully "
            f"successful just because a file was produced."
        )

    return build_excel(equipment_id, pm_results)


async def generate_with_filename(equipment_id: str, company_id: str) -> tuple:
    """
    Preferred entry point for API endpoints: does everything generate()
    does, and also returns a real, human-readable filename alongside the
    bytes -- e.g. "PM_Strategy_P185WJD_Compressor_1_A102_2026-08-05.xlsx"
    instead of whatever generic name a browser falls back to when a
    Content-Disposition header isn't set.

    This exists so the filename can't be forgotten as a separate step --
    generate() itself still returns bytes only and is unchanged, for
    anything already calling it directly.

    Returns: (excel_bytes: bytes, filename: str)
    """
    excel_bytes = await generate(equipment_id, company_id)
    equipment_info = fetch_equipment_info(equipment_id, company_id)
    filename = build_output_filename(equipment_id, equipment_info)
    return excel_bytes, filename